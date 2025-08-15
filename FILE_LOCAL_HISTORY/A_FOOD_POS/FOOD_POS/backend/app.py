# -*- coding: utf-8 -*-
"""
Flask API Server สำหรับระบบ POS ร้านอาหาร
"""

from flask import Flask, request, jsonify, render_template, send_from_directory, send_file, session, redirect, url_for
from flask_cors import CORS
import os
import json
import uuid
import qrcode
import glob
import sqlite3
from datetime import datetime
import pytz
from werkzeug.utils import secure_filename
from PIL import Image
from database import DatabaseManager
from models import *
from utils.qr_generator import QRGenerator
from utils.promptpay import PromptPayGenerator
from utils.google_sheets import GoogleSheetsManager

# เพิ่ม logging สำหรับ debug
import logging
logging.basicConfig(level=logging.DEBUG)

# ฟังก์ชันสำหรับจัดการเวลาท้องถิ่นของไทย
def get_thai_datetime():
    """ได้รับเวลาปัจจุบันในโซนเวลาไทย"""
    thai_tz = pytz.timezone('Asia/Bangkok')
    return datetime.now(thai_tz)

def get_thai_datetime_iso():
    """ได้รับเวลาปัจจุบันในโซนเวลาไทยในรูปแบบ ISO"""
    return get_thai_datetime().isoformat()

print("[STARTUP] Loading Flask app from backend/app.py")

app = Flask(__name__, 
            template_folder='../frontend',
            static_folder='../frontend')
CORS(app)

# Set secret key for session management
app.secret_key = 'your-secret-key-change-this-in-production'

# Set Flask app logger to DEBUG level
app.logger.setLevel(logging.DEBUG)

print("[STARTUP] Flask app initialized successfully")

# เพิ่ม middleware สำหรับ debug requests
@app.before_request
def log_request_info():
    print(f"[MIDDLEWARE] {request.method} {request.url}")
    print(f"[MIDDLEWARE] PATH: {request.path}")
    print(f"[MIDDLEWARE] ARGS: {request.args}")
    print(f"[MIDDLEWARE] ALL REQUESTS: {request.method} {request.path}")
    if 'payment-complete' in request.path:
        print(f"[PAYMENT-COMPLETE] Request detected: {request.method} {request.path}")
        print(f"[PAYMENT-COMPLETE] URGENT: This is a payment request!")
    if request.method == 'POST':
        print(f"[POST-REQUEST] POST request detected: {request.path}")
    app.logger.debug(f"DEBUG REQUEST: {request.method} {request.url}")
    app.logger.debug(f"DEBUG REQUEST PATH: {request.path}")
    app.logger.debug(f"DEBUG REQUEST ARGS: {request.args}")

@app.after_request
def after_request(response):
    """Add cache control headers to prevent browser caching"""
    # ป้องกันการ cache สำหรับ API endpoints
    if request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# Initialize components
# ใช้ absolute path เพื่อให้แน่ใจว่าจะเชื่อมต่อกับฐานข้อมูลที่ถูกต้อง
db_path = os.path.join(os.path.dirname(__file__), "..", "pos_database.db")
db = DatabaseManager(db_path)
qr_gen = QRGenerator()
promptpay_gen = PromptPayGenerator()
sheets_manager = GoogleSheetsManager()

# Global variables
active_sessions = {}  # เก็บ session ที่ active

# ==================== ROUTE PROTECTION ====================

def require_login(f):
    """Decorator สำหรับป้องกัน route ที่ต้อง login"""
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# กำหนดค่าสำหรับการอัปโหลดไฟล์
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'frontend', 'images', 'menu')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'svg'}

# สร้างโฟลเดอร์ถ้ายังไม่มี
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

@app.route('/')
def index():
    """หน้าแรก - แสดงสถานะระบบ"""
    return render_template('index.html')

@app.route('/order')
def order_page():
    """หน้าสั่งอาหารสำหรับลูกค้า"""
    table_id = request.args.get('table')
    session_id = request.args.get('session')
    
    if not table_id:
        return "ไม่พบหมายเลขโต๊ะ", 400
    
    # ตรวจสอบสถานะโต๊ะ
    tables = db.get_all_tables()
    table = next((t for t in tables if t['table_id'] == int(table_id)), None)
    
    if not table:
        return "ไม่พบโต๊ะที่ระบุ", 404
    
    # ตรวจสอบว่าโต๊ะพร้อมใช้งานหรือไม่
    if table['status'] == 'checkout':
        return render_template('error.html', 
                             error_title="QR Code หมดอายุ",
                             error_message="QR Code นี้ไม่สามารถใช้งานได้แล้ว เนื่องจากโต๊ะได้ทำการเช็คบิลแล้ว กรุณาติดต่อพนักงานเพื่อขอ QR Code ใหม่")
    
    # ถ้ามี session_id ให้ตรวจสอบว่าตรงกับโต๊ะหรือไม่
    if session_id and table.get('session_id') and table['session_id'] != session_id:
        return render_template('error.html',
                             error_title="QR Code ไม่ถูกต้อง", 
                             error_message="QR Code นี้ไม่สามารถใช้งานได้ กรุณาติดต่อพนักงานเพื่อขอ QR Code ใหม่")
    
    # Debug: ตรวจสอบไฟล์ template
    template_path = os.path.join(app.template_folder, 'order.html')
    print(f"[DEBUG] Template path: {template_path}")
    print(f"[DEBUG] Template exists: {os.path.exists(template_path)}")
    if os.path.exists(template_path):
        with open(template_path, 'r', encoding='utf-8') as f:
            content = f.read()
            print(f"[DEBUG] Template file size: {len(content)} characters")
            print(f"[DEBUG] Contains callStaffModal: {'callStaffModal' in content}")
    
    return render_template('order.html', table_id=table_id, session_id=session_id)

@app.route('/admin')
@require_login
def admin_page():
    """หน้า Admin Panel สำหรับจัดการระบบ"""
    return render_template('admin.html')

@app.route('/error')
def error_page():
    """หน้าแสดงข้อผิดพลาด"""
    error_title = request.args.get('title', 'เกิดข้อผิดพลาด')
    error_message = request.args.get('message', 'ระบบมีปัญหา กรุณาลองใหม่อีกครั้ง')
    return render_template('error.html', error_title=error_title, error_message=error_message)

@app.route('/test-promptpay')
def test_promptpay_page():
    """หน้าทดสอบการตั้งค่า PromptPay"""
    return render_template('test_promptpay.html')

# === API Endpoints ===

@app.route('/api/settings/promptpay', methods=['GET'])
def get_promptpay_settings():
    """ดึงการตั้งค่า PromptPay ปัจจุบัน"""
    try:
        promptpay_type = db.get_config('promptpay_type') or 'phone'
        promptpay_value = db.get_config('promptpay_value') or ''
        
        return jsonify({
            'success': True,
            'type': promptpay_type,
            'value': promptpay_value
        })
    except Exception as e:
        app.logger.error(f"Error getting PromptPay settings: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/settings/promptpay', methods=['POST'])
def save_promptpay_settings():
    """บันทึกการตั้งค่า PromptPay"""
    try:
        data = request.get_json()
        pp_type = data.get('type')
        pp_value = data.get('value')

        if not pp_type or not pp_value:
            return jsonify({'success': False, 'message': 'ข้อมูลไม่ครบถ้วน'}), 400

        db.save_setting('promptpay_type', pp_type)
        db.save_setting('promptpay_value', pp_value)

        return jsonify({'success': True, 'message': 'บันทึกข้อมูล PromptPay สำเร็จ'})
    except Exception as e:
        app.logger.error(f"Error saving PromptPay settings: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/promptpay/generate-qr', methods=['POST'])
def generate_promptpay_qr():
    """สร้าง PromptPay QR Code โดยใช้การตั้งค่าจากระบบ"""
    try:
        data = request.get_json()
        amount = float(data.get('amount', 0))
        
        if amount <= 0:
            return jsonify({'success': False, 'message': 'จำนวนเงินต้องมากกว่า 0'}), 400
        
        # ดึงการตั้งค่า PromptPay จากระบบ
        promptpay_type = db.get_config('promptpay_type') or 'phone'
        promptpay_value = db.get_config('promptpay_value')
        
        if not promptpay_value:
            return jsonify({'success': False, 'message': 'กรุณาตั้งค่า PromptPay ในหน้าการตั้งค่าก่อน'}), 400
        
        # สร้าง PromptPay QR Code
        qr_code_base64 = promptpay_gen.generate_qr(promptpay_value, amount, promptpay_type)
        
        if not qr_code_base64:
            return jsonify({'success': False, 'message': 'ไม่สามารถสร้าง QR Code ได้'}), 500
        
        # แปลง base64 data URL เป็น base64 string เฉพาะ
        if qr_code_base64.startswith('data:image/png;base64,'):
            qr_code_base64 = qr_code_base64.replace('data:image/png;base64,', '')
            
        return jsonify({
            'success': True, 
            'qr_code': qr_code_base64,
            'promptpay_id': promptpay_value,
            'promptpay_type': promptpay_type,
            'amount': amount
        })
        
    except Exception as e:
        app.logger.error(f"Error generating PromptPay QR: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/promptpay/current-settings', methods=['GET'])
def get_current_promptpay_settings():
    """ดึงการตั้งค่า PromptPay ปัจจุบันสำหรับการตรวจสอบ"""
    try:
        promptpay_type = db.get_config('promptpay_type') or 'phone'
        promptpay_value = db.get_config('promptpay_value') or '0906016218'
        
        return jsonify({
            'success': True,
            'current_type': promptpay_type,
            'current_value': promptpay_value,
            'message': f'ปัจจุบันใช้ {promptpay_type}: {promptpay_value}'
        })
    except Exception as e:
        app.logger.error(f"Error getting current PromptPay settings: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/settings/sheets', methods=['POST'])
def save_google_sheets_settings():
    """บันทึกการตั้งค่า Google Sheets"""
    try:
        sheet_id = request.form.get('sheet_id')
        sheet_name = request.form.get('sheet_name')
        credentials_file = request.files.get('credentials')
        
        if not sheet_id or not sheet_name:
            return jsonify({'success': False, 'error': 'กรุณากรอกข้อมูลที่จำเป็น'}), 400
        
        # บันทึกไฟล์ credentials ถ้ามี
        if credentials_file:
            credentials_filename = secure_filename('credentials.json')
            credentials_path = os.path.join(os.getcwd(), credentials_filename)
            credentials_file.save(credentials_path)
        
        # อัปเดต config
        config = {
            'enabled': True,
            'spreadsheet_id': sheet_id,
            'sheet_names': {
                'orders': 'Orders',
                'order_items': 'Order_Items',
                'daily_summary': 'Daily_Summary'
            }
        }
        
        config_path = os.path.join(os.getcwd(), 'google_sheets_config.json')
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        
        # รีโหลด Google Sheets manager
        global sheets_manager
        sheets_manager = GoogleSheetsManager()
        
        return jsonify({
            'success': True,
            'message': 'บันทึกการตั้งค่า Google Sheets สำเร็จ'
        })
        
    except Exception as e:
        app.logger.error(f"Error saving Google Sheets settings: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/settings/sheets/test', methods=['POST'])
def test_google_sheets_connection():
    """ทดสอบการเชื่อมต่อ Google Sheets"""
    try:
        global sheets_manager
        if not sheets_manager:
            sheets_manager = GoogleSheetsManager()
        
        if sheets_manager.test_connection():
            return jsonify({
                'success': True,
                'message': 'เชื่อมต่อ Google Sheets สำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถเชื่อมต่อ Google Sheets ได้'
            }), 500
            
    except Exception as e:
        app.logger.error(f"Error testing Google Sheets connection: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/tables', methods=['GET'])
def get_tables():
    """ดึงข้อมูลโต๊ะทั้งหมด"""
    try:
        tables = db.get_all_tables()
        return jsonify(tables)
    except Exception as e:
        print(f"[ERROR] get_all_tables failed: {str(e)}")
        print(f"[ERROR] Exception type: {type(e).__name__}")
        import traceback
        print(f"[ERROR] Traceback: {traceback.format_exc()}")
        return jsonify([]), 500

@app.route('/api/tables/<int:table_id>', methods=['GET'])
def get_table(table_id):
    """ดึงข้อมูลโต๊ะเดียว"""
    try:
        table = db.get_table(table_id)
        if table:
            return jsonify({
                'success': True,
                'data': table
            })
        else:
            return jsonify({
                'success': False,
                'error': f'ไม่พบโต๊ะหมายเลข {table_id}'
            }), 404
    except Exception as e:
        print(f"[ERROR] Exception in get_all_orders: {str(e)}")
        print(f"[ERROR] Exception type: {type(e).__name__}")
        import traceback
        print(f"[ERROR] Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tables/<int:table_id>/status', methods=['PUT'])
def update_table_status(table_id):
    """อัปเดตสถานะโต๊ะ"""
    try:
        data = request.get_json()
        status = data.get('status')
        session_id = data.get('session_id')
        
        success = db.update_table_status(table_id, status, session_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'อัปเดตสถานะโต๊ะสำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถอัปเดตสถานะโต๊ะได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tables', methods=['POST'])
def add_table():
    """เพิ่มโต๊ะใหม่"""
    try:
        data = request.get_json()
        table_id = data.get('table_id')
        table_name = data.get('table_name')
        
        if not table_id or not table_name:
            return jsonify({
                'success': False,
                'error': 'กรุณาระบุหมายเลขโต๊ะและชื่อโต๊ะ'
            }), 400
        
        # ตรวจสอบว่าโต๊ะมีอยู่แล้วหรือไม่
        existing_table = db.get_table(table_id)
        if existing_table:
            return jsonify({
                'success': False,
                'error': f'โต๊ะหมายเลข {table_id} มีอยู่แล้ว'
            }), 400
        
        success = db.add_table(table_id, table_name)
        
        if success:
            return jsonify({
                'success': True,
                'message': f'เพิ่มโต๊ะ {table_name} สำเร็จ',
                'table': {
                    'table_id': table_id,
                    'table_name': table_name,
                    'status': 'available'
                }
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถเพิ่มโต๊ะได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tables/<int:table_id>', methods=['DELETE'])
def delete_table(table_id):
    """ลบโต๊ะ"""
    try:
        # ตรวจสอบว่าโต๊ะมีอยู่หรือไม่
        table = db.get_table(table_id)
        if not table:
            return jsonify({
                'success': False,
                'error': f'ไม่พบโต๊ะหมายเลข {table_id}'
            }), 404
        
        success = db.delete_table(table_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': f'ลบโต๊ะ {table["table_name"]} สำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถลบโต๊ะได้ อาจมีออเดอร์ที่ยังไม่เสร็จสิ้น'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tables/<int:table_id>/qr', methods=['GET'])
def generate_table_qr(table_id):
    """สร้าง QR Code สำหรับโต๊ะ"""
    try:
        # ตรวจสอบว่าโต๊ะมี session_id อยู่แล้วหรือไม่
        table = db.get_table(table_id)
        
        if table and table.get('session_id') and table.get('status') == 'occupied':
            # ใช้ session_id เดิมถ้าโต๊ะมีเซสชั่นอยู่แล้ว
            session_id = table['session_id']
            print(f"[DEBUG] Using existing session_id: {session_id} for table {table_id}")
        else:
            # สร้าง session_id ใหม่สำหรับลูกค้า
            session_id = str(uuid.uuid4())
            print(f"[DEBUG] Creating new session_id: {session_id} for table {table_id}")
            
            # อัปเดตสถานะโต๊ะเป็น occupied และกำหนด session_id
            db.update_table_status(table_id, 'occupied', session_id)
        
        domain = db.get_config('domain_url') or 'http://localhost:5000'
        qr_url = f"{domain}/order?table={table_id}&session={session_id}"
        
        qr_code = qr_gen.generate_qr(qr_url)
        
        return jsonify({
            'success': True,
            'data': {
                'qr_code': qr_code,
                'url': qr_url,
                'session_id': session_id
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/menu/categories', methods=['GET'])
def get_menu_categories():
    """ดึงหมวดหมู่เมนู"""
    try:
        print("[DEBUG] Getting menu categories...")
        print(f"[DEBUG] db object type: {type(db)}")
        print(f"[DEBUG] db object: {db}")
        print(f"[DEBUG] Calling db.get_menu_categories()...")
        categories = db.get_menu_categories()
        print(f"[DEBUG] get_menu_categories returned: {categories}")
        print(f"[DEBUG] Found {len(categories)} categories")
        if categories:
            print(f"[DEBUG] First category: {categories[0]}")
        return jsonify(categories)
    except Exception as e:
        print(f"[DEBUG] Error getting categories: {e}")
        import traceback
        print(f"[DEBUG] Traceback: {traceback.format_exc()}")
        return jsonify([]), 500



@app.route('/api/menu/items', methods=['GET', 'POST'])
def handle_menu_items():
    """จัดการเมนูอาหาร - GET: ดึงเมนู, POST: เพิ่มเมนู"""
    if request.method == 'GET':
        try:
            category_id = request.args.get('category_id')
            print(f"[DEBUG] get_menu_items called with category_id: {category_id}")
            items = db.get_menu_items(int(category_id) if category_id else None)
            print(f"[DEBUG] Retrieved {len(items)} items from database")
            print(f"[DEBUG] First 2 items: {items[:2] if items else 'No items'}")
            return jsonify({
                'success': True,
                'data': items
            })
        except Exception as e:
            print(f"[ERROR] get_menu_items error: {str(e)}")
            return jsonify({
                'success': False,
                'error': str(e)
            }), 500
    
    elif request.method == 'POST':
        try:
            data = request.get_json()
            item_id = db.add_menu_item(
                data['name'],
                float(data['price']),
                int(data['category_id']),
                data.get('description', ''),
                data.get('image_url'),
                data.get('is_available', True),
                data.get('preparation_time', 15),
                data.get('food_option_type', 'none')
            )
            
            if item_id:
                return jsonify({
                    'success': True,
                    'data': {'item_id': item_id},
                    'message': 'เพิ่มเมนูสำเร็จ'
                })
            else:
                return jsonify({
                    'success': False,
                    'error': 'ไม่สามารถเพิ่มเมนูได้'
                }), 500
        except Exception as e:
            return jsonify({
                'success': False,
                'error': str(e)
            }), 500

@app.route('/api/menu/items/all', methods=['GET'])
def handle_all_menu_items():
    """ดึงเมนูอาหารทั้งหมด รวมถึงรายการที่ไม่พร้อมจำหน่าย (สำหรับหน้าจัดการเมนู)"""
    try:
        category_id = request.args.get('category_id')
        print(f"[DEBUG] get_all_menu_items called with category_id: {category_id}")
        items = db.get_all_menu_items(int(category_id) if category_id else None)
        print(f"[DEBUG] Retrieved {len(items)} items from database (including unavailable)")
        print(f"[DEBUG] First 2 items: {items[:2] if items else 'No items'}")
        return jsonify({
            'success': True,
            'data': items
        })
    except Exception as e:
        print(f"[ERROR] get_all_menu_items error: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/menu/category/<int:category_id>', methods=['GET'])
def get_menu_by_category(category_id):
    """ดึงเมนูอาหารตามหมวดหมู่"""
    try:
        items = db.get_menu_items(category_id)
        return jsonify(items)
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/menu/categories', methods=['POST'])
def add_menu_category():
    """เพิ่มหมวดหมู่เมนู"""
    try:
        data = request.get_json()
        category_id = db.add_menu_category(
            data['name'],
            data.get('description', '')
        )
        
        if category_id:
            return jsonify({
                'success': True,
                'data': {'category_id': category_id},
                'message': 'เพิ่มหมวดหมู่สำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถเพิ่มหมวดหมู่ได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/menu/categories/<int:category_id>', methods=['PUT'])
def update_menu_category(category_id):
    """อัปเดตหมวดหมู่เมนู"""
    try:
        data = request.get_json()
        success = db.update_menu_category(
            category_id,
            data['name'],
            data.get('description', '')
        )
        
        if success:
            return jsonify({
                'success': True,
                'message': 'อัปเดตหมวดหมู่สำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถอัปเดตหมวดหมู่ได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/menu/categories/<int:category_id>', methods=['DELETE'])
def delete_menu_category(category_id):
    """ลบหมวดหมู่เมนู"""
    try:
        success = db.delete_menu_category(category_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'ลบหมวดหมู่สำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถลบหมวดหมู่ได้ อาจมีเมนูในหมวดหมู่นี้อยู่'
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/menu/categories/<int:category_id>/move-up', methods=['POST'])
def move_category_up(category_id):
    """เลื่อนหมวดหมู่ขึ้น"""
    print(f"[DEBUG] move_category_up called with category_id={category_id}")
    app.logger.debug(f"[DEBUG] move_category_up called with category_id={category_id}")
    try:
        success = db.move_category_up(category_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'เลื่อนหมวดหมู่ขึ้นสำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถเลื่อนหมวดหมู่ขึ้นได้'
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/menu/categories/<int:category_id>/move-down', methods=['POST'])
def move_category_down(category_id):
    """เลื่อนหมวดหมู่ลง"""
    print(f"[DEBUG] move_category_down called with category_id={category_id}")
    app.logger.debug(f"[DEBUG] move_category_down called with category_id={category_id}")
    try:
        success = db.move_category_down(category_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'เลื่อนหมวดหมู่ลงสำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถเลื่อนหมวดหมู่ลงได้'
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/menu/items/<int:item_id>', methods=['PUT'])
def update_menu_item(item_id):
    """อัปเดตเมนูอาหาร"""
    try:
        data = request.get_json()
        print(f"DEBUG: Received data for item {item_id}: {data}")
        print(f"DEBUG: food_option_type value: {data.get('food_option_type', 'none')}")
        
        success = db.update_menu_item(
            item_id,
            data['name'],
            float(data['price']),
            int(data['category_id']),
            data.get('description', ''),
            data.get('image_url'),
            data.get('is_available', True),
            data.get('preparation_time', 15),
            data.get('food_option_type', 'none')
        )
        
        print(f"DEBUG: Update result: {success}")
        
        if success:
            return jsonify({
                'success': True,
                'message': 'อัปเดตเมนูสำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถอัปเดตเมนูได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/menu/items/<int:item_id>', methods=['DELETE'])
def delete_menu_item(item_id):
    """ลบเมนูอาหาร"""
    try:
        success = db.delete_menu_item(item_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'ลบเมนูสำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถลบเมนูได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# === Option Values Management ===
@app.route('/api/option-values', methods=['GET'])
def get_option_values():
    """ดึงค่าตัวเลือกทั้งหมดหรือตามประเภท"""
    try:
        option_type = request.args.get('option_type')
        app.logger.info(f"DEBUG API: Getting option values for type: {option_type}")
        option_values = db.get_option_values(option_type)
        app.logger.info(f"DEBUG API: Retrieved {len(option_values)} option values")
        for option in option_values:
            app.logger.info(f"DEBUG API: Option {option['option_value_id']}: name='{option['name']}', additional_price={option['additional_price']}")
        return jsonify({
            'success': True,
            'data': option_values
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/option-values', methods=['POST'])
def add_option_value():
    """เพิ่มค่าตัวเลือกใหม่"""
    try:
        data = request.get_json()
        success = db.add_option_value(
            data['option_type'],
            data['name'],
            data.get('additional_price', 0),
            data.get('is_default', False),
            data.get('sort_order', 0)
        )
        
        if success:
            return jsonify({
                'success': True,
                'message': 'เพิ่มค่าตัวเลือกสำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถเพิ่มค่าตัวเลือกได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/option-values/<int:option_value_id>', methods=['PUT'])
def update_option_value(option_value_id):
    """อัปเดตค่าตัวเลือก"""
    try:
        data = request.get_json()
        print(f"DEBUG: Updating option_value_id {option_value_id} with data: {data}")
        success = db.update_option_value(
            option_value_id,
            data.get('name'),
            data.get('additional_price'),
            data.get('is_default'),
            data.get('sort_order')
        )
        print(f"DEBUG: Update result: {success}")
        
        if success:
            return jsonify({
                'success': True,
                'message': 'อัปเดตค่าตัวเลือกสำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถอัปเดตค่าตัวเลือกได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/option-values/<int:option_value_id>', methods=['DELETE'])
def delete_option_value(option_value_id):
    """ลบค่าตัวเลือก"""
    try:
        success = db.delete_option_value(option_value_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'ลบค่าตัวเลือกสำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถลบค่าตัวเลือกได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/option-values/set-default', methods=['POST'])
def set_default_option_value():
    """ตั้งค่าเริ่มต้นสำหรับตัวเลือก"""
    try:
        data = request.get_json()
        option_type = data['option_type']
        default_option_id = data['default_option_id']
        
        success = db.set_default_option_value(option_type, default_option_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'ตั้งค่าเริ่มต้นสำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถตั้งค่าเริ่มต้นได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# === Option Types Management ===
@app.route('/api/option-types', methods=['GET'])
def get_option_types():
    """ดึงประเภทตัวเลือกทั้งหมด"""
    try:
        option_types = db.get_option_types()
        return jsonify({
            'success': True,
            'data': option_types
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/option-types', methods=['POST'])
def add_option_type():
    """เพิ่มประเภทตัวเลือกใหม่"""
    try:
        data = request.get_json()
        
        # ตรวจสอบข้อมูลที่จำเป็น
        if not data.get('name') or not data.get('key'):
            return jsonify({
                'success': False,
                'error': 'กรุณากรอกชื่อและรหัสประเภทตัวเลือก'
            }), 400
        
        # ตรวจสอบรูปแบบของ key
        import re
        if not re.match(r'^[a-z_]+$', data['key']):
            return jsonify({
                'success': False,
                'error': 'รหัสประเภทต้องเป็นตัวอักษรภาษาอังกฤษพิมพ์เล็กและขีดล่างเท่านั้น'
            }), 400
        
        success = db.add_option_type(
            data['name'],
            data['key'],
            data.get('description', ''),
            data.get('is_active', True)
        )
        
        if success:
            return jsonify({
                'success': True,
                'message': 'เพิ่มประเภทตัวเลือกสำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถเพิ่มประเภทตัวเลือกได้ อาจมีรหัสประเภทนี้อยู่แล้ว'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/option-types/<int:option_type_id>', methods=['PUT'])
def update_option_type(option_type_id):
    """อัปเดตประเภทตัวเลือก"""
    try:
        data = request.get_json()
        success = db.update_option_type(
            option_type_id,
            data.get('name'),
            data.get('description'),
            data.get('is_active')
        )
        
        if success:
            return jsonify({
                'success': True,
                'message': 'อัปเดตประเภทตัวเลือกสำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถอัปเดตประเภทตัวเลือกได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/option-types/<int:option_type_id>', methods=['DELETE'])
def delete_option_type(option_type_id):
    """ลบประเภทตัวเลือก"""
    try:
        success = db.delete_option_type(option_type_id)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'ลบประเภทตัวเลือกสำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถลบประเภทตัวเลือกได้ อาจมีข้อมูลที่เกี่ยวข้องอยู่'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/orders', methods=['POST'])
def create_order():
    """สร้างออเดอร์ใหม่"""
    try:
        data = request.get_json()
        print(f"[DEBUG] create_order: Received data: {data}")
        table_id = int(data['table_id'])
        items = data['items']  # [{'item_id': 1, 'quantity': 2, 'customer_request': ''}]
        print(f"[DEBUG] create_order: table_id={table_id}, items={items}")
        
        # สร้าง session_id ใหม่หรือใช้ที่มีอยู่
        session_id = data.get('session_id') or str(uuid.uuid4())
        
        # ตรวจสอบว่ามี order active หรือ pending อยู่หรือไม่
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT order_id FROM orders 
            WHERE table_id = ? AND session_id = ? AND status IN ('active', 'pending')
            ORDER BY created_at DESC LIMIT 1
        ''', (table_id, session_id))
        existing_order = cursor.fetchone()
        conn.close()
        
        if existing_order:
            # ใช้ order_id ที่มีอยู่และยังไม่เสร็จสิ้น
            order_id = existing_order[0]
        else:
            # สร้าง order ใหม่
            order_id = db.create_order(table_id, session_id)
            if not order_id:
                raise Exception("ไม่สามารถสร้างออเดอร์ได้")
        
        # เพิ่มรายการอาหาร
        print(f"[DEBUG] create_order: Processing {len(items)} items for order_id {order_id}")
        print(f"[DEBUG] create_order: Full request data: {data}")
        for item in items:
            print(f"[DEBUG] create_order: Processing item: {item}")
            print(f"[DEBUG] create_order: Item keys: {list(item.keys())}")
            print(f"[DEBUG] create_order: selected_option value: {item.get('selected_option')}")
            print(f"[DEBUG] create_order: note value: {item.get('note')}")
            print(f"[DEBUG] create_order: customer_request value: {item.get('customer_request')}")
            
            # ดึงข้อมูลเมนูจากฐานข้อมูลเสมอเพื่อใช้ในการแสดงข้อผิดพลาด
            menu_items = db.get_menu_items()
            menu_item = next((m for m in menu_items if m['item_id'] == item['item_id']), None)
            
            if not menu_item:
                print(f"[ERROR] create_order: Menu item with ID {item['item_id']} not found!")
                continue
            
            # ใช้ราคาที่ frontend ส่งมา (รวมตัวเลือกพิเศษแล้ว) หรือดึงจากฐานข้อมูลถ้าไม่มี
            if 'price' in item and item['price'] is not None:
                unit_price = float(item['price'])
                print(f"[DEBUG] create_order: Using price from frontend: {unit_price}")
            else:
                unit_price = float(menu_item['price'])
                print(f"[DEBUG] create_order: Using price from database: {unit_price}")
            
            # รับค่า selected_option และ notes จาก frontend และรวมเป็น customer_request
            selected_option = item.get('selected_option', '')
            notes = item.get('notes', '') or item.get('note', '')  # รองรับทั้ง 'notes' และ 'note'
            
            print(f"[DEBUG] create_order: Raw item data: {item}")
            print(f"[DEBUG] create_order: selected_option='{selected_option}', notes='{notes}'")
            
            # รวม selected_option และ notes เป็น customer_request
            customer_request_parts = []
            if selected_option:
                customer_request_parts.append(selected_option)
                print(f"[DEBUG] create_order: Added selected_option to parts: '{selected_option}'")
            if notes:
                customer_request_parts.append(notes)
                print(f"[DEBUG] create_order: Added notes to parts: '{notes}'")
            
            customer_request = ' | '.join(customer_request_parts) if customer_request_parts else ''
            print(f"[DEBUG] create_order: Final customer_request: '{customer_request}'")
                
            print(f"[DEBUG] create_order: Adding order item - order_id: {order_id}, item_id: {item['item_id']}, quantity: {item['quantity']}, price: {unit_price}, customer_request: '{customer_request}'")
            
            print(f"[DEBUG] create_order: About to call add_order_item with customer_request: '{customer_request}'")
            success = db.add_order_item(
                order_id,
                item['item_id'],
                item['quantity'],
                unit_price,
                customer_request
            )
            print(f"[DEBUG] create_order: add_order_item returned: {success}")
            
            # ตรวจสอบทันทีหลังจากเพิ่ม
            import sqlite3
            try:
                conn = sqlite3.connect('pos_database.db')
                cursor = conn.cursor()
                cursor.execute('SELECT customer_request FROM order_items WHERE order_id = ? ORDER BY order_item_id DESC LIMIT 1', (order_id,))
                result = cursor.fetchone()
                if result:
                    print(f"[DEBUG] create_order: Verified in DB - customer_request: '{result[0]}'")
                else:
                    print(f"[DEBUG] create_order: No item found in DB for order_id: {order_id}")
                conn.close()
            except Exception as e:
                print(f"[DEBUG] create_order: DB verification error: {e}")
            
            print(f"[DEBUG] create_order: add_order_item result: {success}")
            
            if not success:
                raise Exception(f"ไม่สามารถเพิ่มรายการ {menu_item['name']} ได้")
        
        # อัปเดตสถานะโต๊ะ
        db.update_table_status(table_id, 'occupied', session_id)
        
        # บันทึกการแจ้งเตือนสำหรับการสั่งอาหาร
        try:
            table_info = db.get_table_by_id(table_id)
            table_name = table_info['table_name'] if table_info else f'โต๊ะ {table_id}'
            
            notification_data = {
                'table_id': table_id,
                'message': f'{table_name} สั่งอาหารใหม่ ({len(items)} รายการ)',
                'type': 'order',
                'is_read': False
            }
            db.save_notification(notification_data)
        except Exception as e:
            print(f"Error saving order notification: {e}")
        
        # ซิงค์ข้อมูลไปยัง Google Sheets แบบ background (ไม่ให้ผู้ใช้รอ)
        print(f"[DEBUG] เริ่มซิงค์ Google Sheets สำหรับออเดอร์ {order_id} แบบ background")
        try:
            from .new_google_sheets_sync import sync_order_to_new_format
            import threading
            
            def sync_to_sheets_background():
                try:
                    # ดึงข้อมูลออเดอร์และรายการอาหาร
                    order_data = {
                        'order_id': order_id,
                        'table_id': table_id,
                        'session_id': session_id,
                        'status': 'รอดำเนินการ',
                        'created_at': get_thai_datetime().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    # ดึงรายการอาหารในออเดอร์
                    conn = db.get_connection()
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT mi.name as item_name, oi.quantity, oi.unit_price, 
                               oi.total_price, oi.customer_request
                        FROM order_items oi
                        JOIN menu_items mi ON oi.item_id = mi.item_id
                        WHERE oi.order_id = ?
                    ''', (order_id,))
                    
                    order_items = []
                    for row in cursor.fetchall():
                        order_items.append({
                            'item_name': row[0],
                            'quantity': row[1],
                            'unit_price': row[2],
                            'total_price': row[3],
                            'customer_request': row[4] or '',
                            'special_options': ''  # สำหรับความเข้ากันได้
                        })
                    conn.close()
                    
                    # ซิงค์ไปยัง Google Sheets
                    sync_success = sync_order_to_new_format(order_data, order_items)
                    if sync_success:
                        print(f"[Google Sheets] ซิงค์ออเดอร์ {order_id} สำเร็จ")
                    else:
                        print(f"[Google Sheets] ไม่สามารถซิงค์ออเดอร์ {order_id} ได้")
                        
                except Exception as sheets_error:
                    print(f"[Google Sheets] เกิดข้อผิดพลาดในการซิงค์: {sheets_error}")
            
            # เรียกใช้ sync ใน background thread
            sync_thread = threading.Thread(target=sync_to_sheets_background)
            sync_thread.daemon = True
            sync_thread.start()
            print(f"[DEBUG] เริ่ม background sync thread สำหรับออเดอร์ {order_id}")
                
        except Exception as e:
            print(f"[Google Sheets] เกิดข้อผิดพลาดในการเริ่ม background sync: {e}")
        
        return jsonify({
            'success': True,
            'data': {
                'order_id': order_id,
                'session_id': session_id
            },
            'message': 'สั่งอาหารสำเร็จ'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/orders', methods=['GET'])
def get_all_orders():
    """ดึงรายการออเดอร์ทั้งหมด"""
    try:
        print("[DEBUG] Starting get_all_orders function")
        # ดึงออเดอร์ทั้งหมดจากฐานข้อมูล
        conn = db.get_connection()
        print("[DEBUG] Database connection established")
        cursor = conn.cursor()
        
        # ดึงข้อมูลออเดอร์ทั้งหมด (เฉพาะออเดอร์ที่มี session_id ตรงกับโต๊ะที่ยังมี session_id อยู่)
        print("[DEBUG] Executing main orders query")
        cursor.execute("""
            SELECT DISTINCT o.order_id, o.table_id, o.session_id, o.status, 
                   o.created_at, t.table_name
            FROM orders o
            JOIN tables t ON o.table_id = t.table_id
            WHERE o.status IN ('pending', 'accepted', 'completed', 'rejected', 'active')
            AND t.session_id IS NOT NULL
            AND o.session_id = t.session_id
            ORDER BY o.created_at DESC
        """)
        print("[DEBUG] Main query executed successfully")
        
        orders = []
        order_rows = cursor.fetchall()
        print(f"[DEBUG] Found {len(order_rows)} orders")
        
        for i, order_row in enumerate(order_rows):
            print(f"[DEBUG] Processing order {i+1}/{len(order_rows)}: {order_row}")
            order_id = order_row[0]
            
            # ดึงรายการอาหารของแต่ละออเดอร์ (รวม total_price ที่คำนวณแล้วจาก frontend)
            # ใช้ customer_request (แก้ไขจาก special_request)
            print(f"[DEBUG] Executing order items query for order_id: {order_id}")
            try:
                cursor.execute("""
                    SELECT mi.name, oi.quantity, oi.unit_price, oi.customer_request, oi.total_price, oi.order_item_id, oi.status
                    FROM order_items oi
                    JOIN menu_items mi ON oi.item_id = mi.item_id
                    WHERE oi.order_id = ?
                """, (order_id,))
                print(f"[DEBUG] Order items query executed for order_id: {order_id}")
                
                items = []
                total_amount = 0
                for item_row in cursor.fetchall():
                    # ใช้ total_price ที่ frontend คำนวณแล้วรวม special options
                    item_total = item_row[4] if item_row[4] is not None else (item_row[1] * item_row[2])
                    total_amount += item_total
                    items.append({
                        'name': item_row[0],
                        'quantity': item_row[1],
                        'price': item_row[2],
                        'total_price': item_total,
                        'customer_request': item_row[3] if item_row[3] else '',
                        'order_item_id': item_row[5],
                        'status': item_row[6] if item_row[6] else 'pending'
                    })
            except Exception as item_error:
                print(f"[ERROR] Error processing order items for order_id {order_id}: {item_error}")
                items = []
                total_amount = 0
            
            orders.append({
                'order_id': order_row[0],
                'table_id': order_row[1],
                'table_name': order_row[5],
                'session_id': order_row[2],
                'status': order_row[3],
                'created_at': order_row[4],
                'items': items,
                'total_amount': total_amount
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': orders
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/debug/orders', methods=['GET'])
def debug_orders():
    """ตรวจสอบสถานะออเดอร์ในฐานข้อมูล"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT order_id, table_id, status, created_at 
            FROM orders 
            ORDER BY created_at DESC 
            LIMIT 10
        """)
        
        orders = []
        for row in cursor.fetchall():
            orders.append({
                'order_id': row[0],
                'table_id': row[1], 
                'status': row[2],
                'created_at': row[3]
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': orders
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500



@app.route('/api/tables/<int:table_id>/orders', methods=['GET'])
def get_table_order_details(table_id):
    """ดึงประวัติการสั่งอาหารของโต๊ะ (แสดงรายการอาหารและสรุปราคา)"""
    print(f"[FUNCTION] get_table_order_details called with table_id={table_id}")
    app.logger.debug("=== FUNCTION get_table_order_details CALLED ===")
    app.logger.debug(f"Table ID: {table_id}")
    try:
        session_id = request.args.get('session_id')
        print(f"[DEBUG] get_table_order_details: Received session_id from request: {session_id}")
        app.logger.debug(f"DEBUG API: get_table_order_details called with table_id={table_id}, session_id={session_id}")
        
        # ดึงข้อมูลโต๊ะ
        table = db.get_table(table_id)
        if not table:
            return jsonify([]), 404
            
        print(f"[DEBUG] get_table_order_details: Table session_id from database: {table.get('session_id')}")
        
        # ดึงออเดอร์ทั้งหมดของโต๊ะ โดยใช้ session_id เสมอ แม้จะเป็น None
        # ถ้า session_id เป็น None แต่ได้รับจาก query parameter ให้ใช้ session_id จากตาราง
        if session_id is None and 'session_id' in request.args:
            # ถ้ามีการส่ง session_id=null หรือ session_id= มาจาก frontend
            # ให้ใช้ session_id จากตาราง
            session_id = table.get('session_id')
            print(f"[DEBUG] get_table_order_details: Using table's session_id instead: {session_id}")
        
        orders = db.get_table_orders(table_id, session_id)
        print(f"[DEBUG] get_table_order_details: Using session_id for query: {session_id}")
        app.logger.debug(f"DEBUG API: orders returned from db.get_table_orders: {len(orders)} items")
        
        if not orders:
            # ส่งกลับโครงสร้างข้อมูลที่ถูกต้องแม้ไม่มีออเดอร์
            response_data = {
                'table_id': table_id,
                'table_name': f'โต๊ะ {table_id}',
                'status': table.get('status', 'unknown'),
                'session_id': table.get('session_id'),
                'orders': [],
                'total_amount': 0,
                'order_count': 0
            }
            return jsonify({
                'success': True,
                'data': response_data
            })
        
        # จัดกลุ่มรายการตามชื่ออาหารและ special request
        grouped_orders = {}
        total_amount = 0
        
        for order in orders:
            item_name = order['item_name']
            item_id = order['item_id']
            customer_request = order.get('customer_request', '') or ''
            quantity = order['quantity']
            unit_price = order['unit_price']
            item_total = order['total_price']
            
            # สร้าง key สำหรับจัดกลุ่ม (ชื่ออาหาร + customer request)
            group_key = f"{item_name}|{customer_request}|{order['item_status']}"
            
            if group_key in grouped_orders:
                # รวมจำนวนและราคา
                grouped_orders[group_key]['quantity'] += quantity
                grouped_orders[group_key]['total_price'] += item_total
            else:
                grouped_orders[group_key] = {
                    'item_id': item_id,
                    'name': item_name,
                    'quantity': quantity,
                    'unit_price': unit_price,
                    'total_price': item_total,
                    'customer_request': customer_request,
                    'item_status': order['item_status']
                }
            
            # ไม่รวมราคาของรายการที่ปฏิเสธในยอดรวม
            if order['item_status'] != 'rejected':
                total_amount += item_total
        
        # แปลงเป็น list
        menu_items = list(grouped_orders.values())
        
        # ดึงข้อมูลโต๊ะ
        table_info = db.get_table(table_id)
        
        # ดึง order_id จากออเดอร์แรกที่มีอยู่
        order_id = orders[0]['order_id'] if orders else None
        
        # สร้างโครงสร้างข้อมูลสำหรับส่งกลับ
        response_data = {
            'table_id': table_id,
            'table_name': f'โต๊ะ {table_id}',
            'status': table_info.get('status', 'unknown') if table_info else 'unknown',
            'session_id': table_info.get('session_id') if table_info else None,
            'order_id': order_id,  # เพิ่ม order_id
            'orders': [{
                'menu_id': item['item_id'],  # เพิ่ม menu_id
                'menu_name': item['name'],
                'quantity': item['quantity'],
                'price': item['unit_price'],
                'total': item['total_price'],
                'customer_request': item['customer_request'],
                'item_status': item['item_status']
            } for item in menu_items],
            'total_amount': total_amount,
            'order_count': len(menu_items)
        }
        
        app.logger.debug(f"DEBUG API: Final response_data: {len(response_data['orders'])} orders")
        
        return jsonify({
            'success': True,
            'data': response_data
        })
        
    except Exception as e:
        app.logger.error(f"Error in get_table_order_details: {str(e)}")
        return jsonify([]), 500

@app.route('/api/tables/<int:table_id>/call', methods=['POST'])
def call_staff(table_id):
    """เรียกพนักงาน"""
    try:
        # อัปเดตสถานะโต๊ะเป็น calling
        success = db.update_table_status(table_id, 'calling')
        
        if success:
            return jsonify({
                'success': True,
                'message': f'เรียกพนักงานที่โต๊ะ {table_id} สำเร็จ'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถเรียกพนักงานได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tables/<int:table_id>/checkout', methods=['POST'])
def checkout_table(table_id):
    """เช็คบิลโต๊ะ"""
    try:
        data = request.get_json()
        session_id = data.get('session_id')
        
        # ดึงออเดอร์ของโต๊ะ
        orders = db.get_table_orders(table_id, session_id)
        
        if not orders:
            return jsonify({
                'success': False,
                'error': 'ไม่พบออเดอร์สำหรับโต๊ะนี้'
            }), 404
        
        # คำนวณยอดรวม (ไม่รวมรายการที่ถูก reject)
        total_amount = sum(order['total_price'] for order in orders if order.get('status') != 'rejected')
        
        # สร้าง PromptPay QR Code
        promptpay_id = db.get_config('promptpay_id')
        promptpay_qr = ""
        
        if promptpay_id:
            promptpay_qr = promptpay_gen.generate_qr(promptpay_id, total_amount)
        
        # ดึง order_id จากออเดอร์แรก (ถ้ามี)
        order_id = orders[0].get('order_id') if orders else None
        
        # สร้างใบเสร็จ
        receipt_data = {
            'table_id': table_id,
            'session_id': session_id,
            'order_id': order_id,
            'orders': orders,
            'total_amount': total_amount,
            'promptpay_qr': promptpay_qr,
            'created_at': get_thai_datetime_iso()
        }
        
        # หมายเหตุ: ไม่ปิดออเดอร์ในขั้นตอนเช็คบิล เพื่อให้สถานะของรายการยังคงเป็นสถานะจริง
        # ออเดอร์จะถูกปิดเมื่อชำระเงินเสร็จสิ้นแล้ว
        
        # อัปเดตสถานะโต๊ะเป็นรอชำระเงิน แต่ยังคงใช้ session_id เดิม
        db.update_table_status(table_id, 'waiting_payment', session_id)
        
        # ส่งข้อมูลไป Google Sheets (ถ้าตั้งค่าไว้)
        try:
            sheets_manager.send_sales_data(receipt_data)
        except Exception as e:
            print(f"Warning: Could not send to Google Sheets: {e}")
        
        return jsonify({
            'success': True,
            'data': receipt_data,
            'message': 'เช็คบิลสำเร็จ'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tables/<int:table_id>/order-summary', methods=['GET'])
def get_table_order_summary(table_id):
    """ดูรายการคำสั่งซื้อของโต๊ะ"""
    try:
        session_id = request.args.get('session_id')
        
        # ดึงข้อมูลโต๊ะ
        table = db.get_table(table_id)
        if not table:
            return jsonify({
                'success': False,
                'error': 'ไม่พบโต๊ะ'
            }), 404
        
        # ใช้ session_id จากตารางถ้าไม่ได้ส่งมา
        if session_id is None:
            session_id = table.get('session_id')
        
        # ดึงรายการคำสั่งซื้อทั้งหมดของโต๊ะ (ทุกสถานะ)
        orders = db.get_table_orders(table_id, session_id)
        
        if not orders:
            return jsonify({
                'success': True,
                'data': {
                    'table_id': table_id,
                    'table_name': f'โต๊ะ {table_id}',
                    'status': table.get('status', 'unknown'),
                    'session_id': table.get('session_id'),
                    'orders': [],
                    'total_amount': 0,
                    'order_count': 0
                }
            })
        
        # จัดกลุ่มรายการตามชื่ออาหารและ special request
        grouped_orders = {}
        total_amount = 0
        
        for order in orders:
            # สร้าง key สำหรับจัดกลุ่ม
            group_key = f"{order['item_name']}_{order.get('customer_request', '')}"
            
            if group_key not in grouped_orders:
                grouped_orders[group_key] = {
                    'item_id': order['item_id'],
                    'name': order['item_name'],
                    'quantity': 0,
                    'unit_price': order['unit_price'],
                    'total_price': 0,
                    'customer_request': order.get('customer_request', ''),
                    'item_status': order.get('item_status', 'pending')
                }
            
            # รวมจำนวนและราคา
            grouped_orders[group_key]['quantity'] += order['quantity']
            grouped_orders[group_key]['total_price'] += order['total_price']
            
            # คำนวณยอดรวมทั้งหมด (ไม่รวมรายการที่ถูกปฏิเสธ)
            if order.get('item_status') != 'rejected':
                total_amount += order['total_price']
        
        # แปลงเป็น list
        menu_items = list(grouped_orders.values())
        
        # สร้างโครงสร้างข้อมูลสำหรับส่งกลับ
        response_data = {
            'table_id': table_id,
            'table_name': f'โต๊ะ {table_id}',
            'status': table.get('status', 'unknown'),
            'session_id': table.get('session_id'),
            'orders': [{
                'menu_id': item['item_id'],
                'menu_name': item['name'],
                'quantity': item['quantity'],
                'unit_price': item['unit_price'],
                'total_price': item['total_price'],
                'customer_request': item['customer_request'],
                'item_status': item['item_status']
            } for item in menu_items],
            'total_amount': total_amount,
            'order_count': len(menu_items)
        }
        
        return jsonify({
            'success': True,
            'data': response_data
        })
        
    except Exception as e:
        app.logger.error(f"Error in get_table_order_summary: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tables/<int:table_id>/update-orders', methods=['POST'])
def update_table_orders(table_id):
    """อัปเดตออเดอร์ของโต๊ะ"""
    conn = None
    try:
        data = request.get_json()
        session_id = data.get('session_id')
        new_orders = data.get('orders', [])
        
        if not session_id:
            return jsonify({
                'success': False,
                'error': 'ไม่พบ session_id'
            }), 400
        
        # ตรวจสอบว่าโต๊ะมีอยู่จริง
        table = db.get_table(table_id)
        if not table:
            return jsonify({
                'success': False,
                'error': 'ไม่พบโต๊ะที่ระบุ'
            }), 404
        
        # ตรวจสอบ session_id
        if table.get('session_id') != session_id:
            return jsonify({
                'success': False,
                'error': 'Session ID ไม่ถูกต้อง'
            }), 403
        
        # หมายเหตุ: ไม่ควรข้ามการประมวลผลแม้ว่า new_orders จะว่าง
        # เพราะอาจเป็นการลบรายการทั้งหมด ซึ่งก็เป็นการเปลี่ยนแปลงที่ถูกต้อง
        # ดังนั้นจะดำเนินการต่อเพื่อล้างออเดอร์ที่มีอยู่
        
        # ใช้ transaction เพื่อให้การอัปเดตเป็น atomic operation
        conn = db.get_connection()
        cursor = conn.cursor()
        
        try:
            # เริ่ม transaction
            cursor.execute('BEGIN TRANSACTION')
            
            # ดึงออเดอร์ปัจจุบันของโต๊ะ
            cursor.execute('''
                SELECT order_id FROM orders 
                WHERE table_id = ? AND session_id = ?
            ''', (table_id, session_id))
            
            existing_order = cursor.fetchone()
            
            if existing_order:
                # ถ้ามีออเดอร์อยู่แล้ว ให้ลบเฉพาะ order_items แล้วเพิ่มใหม่
                order_id = existing_order['order_id']
                
                # ลบ order_items เดิม
                cursor.execute('''
                    DELETE FROM order_items 
                    WHERE order_id = ?
                ''', (order_id,))
                
                # ถ้าไม่มีรายการใหม่ ให้ลบ order record ด้วย
                if not new_orders:
                    print(f"[DEBUG] No new orders, deleting order record {order_id}")
                    cursor.execute('''
                        DELETE FROM orders 
                        WHERE order_id = ?
                    ''', (order_id,))
                    order_id = None
                
            else:
                # ถ้ายังไม่มีออเดอร์และมีรายการใหม่ ให้สร้างใหม่
                if new_orders:
                    cursor.execute('''
                        INSERT INTO orders (table_id, session_id)
                        VALUES (?, ?)
                    ''', (table_id, session_id))
                    order_id = cursor.lastrowid
                    
                    if not order_id:
                        raise Exception('ไม่สามารถสร้างออเดอร์ได้')
                else:
                    # ไม่มีออเดอร์เดิมและไม่มีรายการใหม่ ไม่ต้องทำอะไร
                    order_id = None
            
            # เพิ่มรายการอาหารใหม่ (ถ้ามี)
            if new_orders and order_id:  # ตรวจสอบว่ามีรายการใหม่และมี order_id
                print(f"[DEBUG] Processing {len(new_orders)} new orders")
                for i, order_item in enumerate(new_orders):
                    print(f"[DEBUG] Order item {i}: {order_item}")
                    menu_id = order_item.get('menu_id')
                    quantity = order_item.get('quantity', 1)
                    price = order_item.get('price')
                    customer_request = order_item.get('customer_request', '')
                    status = order_item.get('status', 'pending')
                    
                    if not menu_id or not price:
                        print(f"[DEBUG] Skipping item {i}: missing menu_id or price")
                        continue
                    
                    total_price = order_item.get('total_price', price * quantity)
                    print(f"[DEBUG] Inserting: order_id={order_id}, menu_id={menu_id}, quantity={quantity}, price={price}, total_price={total_price}, status={status}")
                    cursor.execute('''
                        INSERT INTO order_items (order_id, item_id, quantity, unit_price, total_price, customer_request, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (order_id, menu_id, quantity, price, total_price, customer_request, status))
            # หากไม่มีรายการใหม่ (new_orders ว่าง) order record จะถูกลบไปแล้ว
            # ดังนั้นจะไม่มีข้อมูลออเดอร์เหลืออยู่
            
            # Commit transaction
            conn.commit()
            print(f"Successfully updated orders for table {table_id}, session {session_id}")
            
        except Exception as e:
            # Rollback transaction on error
            conn.rollback()
            raise e
        finally:
            if conn:
                conn.close()
        
        return jsonify({
            'success': True,
            'message': 'อัปเดตออเดอร์สำเร็จ'
        })
        
    except Exception as e:
        print(f"Error updating table orders: {e}")
        if conn:
            try:
                conn.rollback()
                conn.close()
            except:
                pass
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tables/<int:table_id>/payment-complete', methods=['POST'])
def payment_complete(table_id):
    """ชำระเงินเสร็จสิ้น"""
    print(f"[DEBUG] payment_complete function called with table_id: {table_id}")
    try:
        # ดึงข้อมูลโต๊ะก่อนอัปเดตเพื่อเก็บ session_id เดิม
        table = db.get_table(table_id)
        session_id = table.get('session_id') if table else None
        print(f"[DEBUG] payment_complete: Table {table_id} with session_id: {session_id}")
        
        # ใช้ transaction เดียวสำหรับการปิดออเดอร์และอัปเดตสถานะโต๊ะ
        success = db.complete_payment_transaction(table_id, session_id)
        
        if success:
            print(f"[DEBUG] payment_complete: Table {table_id} payment completed successfully")
            return jsonify({
                'success': True,
                'message': f'ชำระเงินโต๊ะ {table_id} เสร็จสิ้น'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถดำเนินการชำระเงินได้'
            }), 400
            
    except Exception as e:
        print(f"[ERROR] payment_complete: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tables/<int:table_id>/session/check', methods=['GET'])
def check_session_status(table_id):
    """ตรวจสอบสถานะเซสชั่นของโต๊ะ"""
    try:
        session_id = request.args.get('session_id')
        
        if not session_id:
            return jsonify({
                'success': False,
                'error': 'ไม่พบ session_id',
                'session_valid': False
            }), 400
        
        # ดึงข้อมูลโต๊ะ
        table = db.get_table(table_id)
        
        if not table:
            return jsonify({
                'success': False,
                'error': 'ไม่พบโต๊ะที่ระบุ',
                'session_valid': False
            }), 404
        
        # ตรวจสอบสถานะเซสชั่น
        current_session_id = table.get('session_id')
        table_status = table.get('status')
        
        # เซสชั่นไม่ถูกต้องหาก:
        # 1. session_id เป็น None (ถูกปิดแล้ว)
        # 2. session_id ไม่ตรงกัน
        # 3. สถานะโต๊ะเป็น checkout
        session_valid = (
            current_session_id is not None and 
            current_session_id == session_id and 
            table_status != 'checkout'
        )
        
        return jsonify({
            'success': True,
            'session_valid': session_valid,
            'table_status': table_status,
            'message': 'ตรวจสอบสถานะเซสชั่นสำเร็จ'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'session_valid': False
        }), 500

@app.route('/api/tables/<int:table_id>/clear', methods=['POST'])
def clear_table(table_id):
    """เคลียร์โต๊ะ"""
    try:
        # ดึงข้อมูลโต๊ะก่อนเคลียร์เพื่อเก็บ session_id เดิม
        table = db.get_table(table_id)
        old_session_id = table.get('session_id') if table else None
        print(f"[DEBUG] clear_table: Clearing table {table_id} with old session_id: {old_session_id}")
        
        # รีเซ็ตสถานะโต๊ะและล้างค่า session_id เป็น None อย่างชัดเจน
        success = db.update_table_status(table_id, 'available', None)
        print(f"[DEBUG] clear_table: Table {table_id} cleared, session_id set to None")
        
        if success:
            return jsonify({
                'success': True,
                'message': f'เคลียร์โต๊ะ {table_id} สำเร็จ',
                'old_session_id': old_session_id  # ส่งค่า session_id เดิมกลับไปเพื่อการตรวจสอบ
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถเคลียร์โต๊ะได้'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tables/<int:table_id>/close-session', methods=['POST'])
def close_table_session(table_id):
    """ปิดเซสชั่นโต๊ะและบันทึกข้อมูล"""
    try:
        data = request.get_json()
        session_id = data.get('session_id')
        
        if not session_id:
            return jsonify({
                'success': False,
                'error': 'ไม่พบ session_id'
            }), 400
        
        # ตรวจสอบว่าโต๊ะมีอยู่จริง
        table = db.get_table(table_id)
        if not table:
            return jsonify({
                'success': False,
                'error': 'ไม่พบโต๊ะที่ระบุ'
            }), 404
        
        # ตรวจสอบ session_id
        if table.get('session_id') != session_id:
            return jsonify({
                'success': False,
                'error': 'Session ID ไม่ถูกต้อง'
            }), 403
        
        # บันทึกข้อมูลออเดอร์ลงในประวัติ (ถ้ามี)
        conn = db.get_connection()
        cursor = conn.cursor()
        
        try:
            # ดึงออเดอร์ทั้งหมดของเซสชั่นนี้
            cursor.execute("""
                SELECT o.order_id, o.table_id, o.session_id, o.status, o.created_at,
                       oi.item_id, oi.quantity, oi.unit_price, oi.total_price, oi.customer_request,
                       mi.name as item_name
                FROM orders o
                LEFT JOIN order_items oi ON o.order_id = oi.order_id
                LEFT JOIN menu_items mi ON oi.item_id = mi.item_id
                WHERE o.table_id = ? AND o.session_id = ?
                ORDER BY o.created_at
            """, (table_id, session_id))
            
            session_orders = cursor.fetchall()
            
            if session_orders:
                # บันทึกลงในตาราง session_history (ถ้ามี)
                # หรือเพียงแค่ลบออเดอร์ที่ยังไม่เสร็จสิ้น
                
                # ลบออเดอร์ที่ยังไม่เสร็จสิ้นออกจากระบบ
                cursor.execute("""
                    DELETE FROM order_items 
                    WHERE order_id IN (
                        SELECT order_id FROM orders 
                        WHERE table_id = ? AND session_id = ? AND status != 'completed'
                    )
                """, (table_id, session_id))
                
                cursor.execute("""
                    DELETE FROM orders 
                    WHERE table_id = ? AND session_id = ? AND status != 'completed'
                """, (table_id, session_id))
            
            # อัปเดตสถานะโต๊ะเป็น available และล้าง session_id
            cursor.execute("""
                UPDATE tables 
                SET status = 'available', session_id = NULL 
                WHERE table_id = ?
            """, (table_id,))
            
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': f'ปิดเซสชั่น {session_id[:8]} ของโต๊ะ {table_id} เรียบร้อยแล้ว'
            })
            
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            conn.close()
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/config', methods=['GET'])
def get_config():
    """ดึงการตั้งค่าระบบ"""
    try:
        config = db.get_all_config()
        return jsonify({
            'success': True,
            'data': config
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/config', methods=['POST'])
def update_config():
    """อัปเดตการตั้งค่าระบบ"""
    try:
        data = request.get_json()
        
        for key, value in data.items():
            db.set_config(key, str(value))
        
        return jsonify({
            'success': True,
            'message': 'อัปเดตการตั้งค่าสำเร็จ'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Static files
@app.route('/api/status', methods=['GET'])
def get_server_status():
    """ตรวจสอบสถานะเซิร์ฟเวอร์"""
    try:
        # ทดสอบการเชื่อมต่อฐานข้อมูล
        db.get_all_tables()
        return jsonify({
            'success': True,
            'status': 'online',
            'message': 'เซิร์ฟเวอร์ทำงานปกติ',
            'timestamp': get_thai_datetime_iso()
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'status': 'error',
            'message': f'เซิร์ฟเวอร์มีปัญหา: {str(e)}',
            'timestamp': get_thai_datetime_iso()
        }), 500

@app.route('/api/tools/generate-qr', methods=['POST'])
def generate_all_qr():
    """สร้าง QR Code สำหรับทุกโต๊ะ"""
    try:
        tables = db.get_all_tables()
        domain = db.get_config('domain_url') or 'http://localhost:5000'
        
        qr_codes = {}
        for table in tables:
            table_id = table['table_id']
            
            # สร้าง session_id ใหม่สำหรับแต่ละโต๊ะ
            session_id = str(uuid.uuid4())
            
            # อัปเดตสถานะโต๊ะเป็น occupied และกำหนด session_id
            db.update_table_status(table_id, 'occupied', session_id)
            
            qr_url = f"{domain}/order?table={table_id}&session={session_id}"
            qr_code = qr_gen.generate_qr(qr_url)
            qr_codes[table_id] = {
                'qr_code': qr_code,
                'url': qr_url,
                'session_id': session_id
            }
        
        return jsonify({
            'success': True,
            'data': qr_codes,
            'message': f'สร้าง QR Code สำหรับ {len(qr_codes)} โต๊ะสำเร็จ'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tools/clear-all-tables', methods=['POST'])
def clear_all_tables():
    """เคลียร์ทุกโต๊ะ"""
    try:
        tables = db.get_all_tables()
        cleared_count = 0
        
        for table in tables:
            if table['status'] != 'available':
                db.update_table_status(table['table_id'], 'available')
                cleared_count += 1
        
        return jsonify({
            'success': True,
            'data': {
                'cleared_count': cleared_count,
                'total_tables': len(tables)
            },
            'message': f'เคลียร์ {cleared_count} โต๊ะสำเร็จ'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tools/export-data', methods=['GET'])
def export_data():
    """ส่งออกข้อมูล"""
    try:
        # ส่งออกข้อมูลทั้งหมด
        data = {
            'tables': db.get_all_tables(),
            'menu_categories': db.get_menu_categories(),
            'menu_items': db.get_menu_items(),
            'orders': db.get_all_orders(),
            'export_timestamp': get_thai_datetime_iso()
        }
        
        return jsonify({
            'success': True,
            'data': data,
            'message': 'ส่งออกข้อมูลสำเร็จ'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/tables/<int:table_id>/qr/print', methods=['GET'])
def print_table_qr(table_id):
    """สร้าง QR Code สำหรับพิมพ์"""
    try:
        # สร้าง session_id ใหม่สำหรับลูกค้า
        session_id = str(uuid.uuid4())
        
        # อัปเดตสถานะโต๊ะเป็น occupied และกำหนด session_id
        db.update_table_status(table_id, 'occupied', session_id)
        
        domain = db.get_config('domain_url') or 'http://localhost:5000'
        qr_url = f"{domain}/order?table={table_id}&session={session_id}"
        
        print(f"[DEBUG] Generating QR code for table {table_id} with URL: {qr_url}")
        
        # สร้าง QR Code ขนาดใหญ่สำหรับพิมพ์
        config = {
            'version': 1,
            'error_correction': qrcode.constants.ERROR_CORRECT_M,
            'box_size': 15,  # ขนาดใหญ่ขึ้นสำหรับพิมพ์
            'border': 4,
        }
        
        qr_code = qr_gen.generate_qr(qr_url, config)
        
        if not qr_code:
            print(f"[ERROR] Failed to generate QR code for table {table_id}")
            raise Exception("Failed to generate QR code")
        
        # ตรวจสอบว่า qr_code มีรูปแบบที่ถูกต้องหรือไม่
        if not qr_code.startswith('data:image/png;base64,'):
            print(f"[ERROR] QR code data format is incorrect: {qr_code[:30]}...")
            raise Exception("QR code data format is incorrect")
            
        print(f"[DEBUG] QR code generated successfully. Data length: {len(qr_code) if qr_code else 0}")
        print(f"[DEBUG] QR code data prefix: {qr_code[:30]}...")
        
        # ตัด prefix 'data:image/png;base64,' ออกเพื่อส่งเฉพาะ base64 string
        # เนื่องจาก frontend จะเติม prefix นี้เอง
        base64_only = qr_code.replace('data:image/png;base64,', '')
        
        return jsonify({
            'success': True,
            'data': {
                'table_id': table_id,
                'qr_code': base64_only,  # ส่งเฉพาะ base64 string โดยไม่มี prefix
                'url': qr_url,
                'session_id': session_id,
                'print_ready': True
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/orders/<int:order_id>/accept', methods=['POST'])
def accept_order(order_id):
    """รับออเดอร์"""
    try:
        success = db.update_order_status(order_id, 'accepted')
        if success:
            return jsonify({
                'success': True,
                'message': 'รับออเดอร์เรียบร้อยแล้ว'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถรับออเดอร์ได้'
            }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/orders/<int:order_id>/reject', methods=['POST'])
def reject_order(order_id):
    """ปฏิเสธออเดอร์และเปลี่ยนราคาเป็น 0"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # เปลี่ยนสถานะออเดอร์เป็น rejected
        cursor.execute('''
            UPDATE orders 
            SET status = 'rejected', updated_at = CURRENT_TIMESTAMP
            WHERE order_id = ?
        ''', (order_id,))
        
        # เปลี่ยนราคาของ order_items เป็น 0
        cursor.execute('''
            UPDATE order_items 
            SET unit_price = 0, total_price = 0
            WHERE order_id = ?
        ''', (order_id,))
        
        # อัปเดต total_amount ของออเดอร์เป็น 0
        cursor.execute('''
            UPDATE orders 
            SET total_amount = 0
            WHERE order_id = ?
        ''', (order_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'ปฏิเสธออเดอร์เรียบร้อยแล้ว ราคาถูกเปลี่ยนเป็น 0'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/orders/<int:order_id>/items', methods=['GET'])
def get_order_items(order_id):
    """ดึงรายการออเดอร์ย่อยพร้อมสถานะ"""
    try:
        items = db.get_order_items_with_status(order_id)
        return jsonify({
            'success': True,
            'data': items
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/order-items/<int:order_item_id>/accept', methods=['POST'])
def accept_order_item(order_item_id):
    """รับรายการออเดอร์ย่อย"""
    try:
        success = db.update_order_item_status(order_item_id, 'accepted')
        if success:
            return jsonify({
                'success': True,
                'message': 'รับรายการเรียบร้อยแล้ว'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถรับรายการได้'
            }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/order-items/<int:order_item_id>/reject', methods=['POST'])
def reject_order_item(order_item_id):
    """ปฏิเสธรายการออเดอร์ย่อย"""
    try:
        success = db.update_order_item_status(order_item_id, 'rejected')
        if success:
            return jsonify({
                'success': True,
                'message': 'ปฏิเสธรายการเรียบร้อยแล้ว'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถปฏิเสธรายการได้'
            }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/order-items/<int:order_item_id>/complete', methods=['POST'])
def complete_order_item(order_item_id):
    """ทำเครื่องหมายรายการออเดอร์ย่อยเสร็จสิ้น"""
    try:
        success = db.update_order_item_status(order_item_id, 'completed')
        if success:
            return jsonify({
                'success': True,
                'message': 'ทำเครื่องหมายรายการเสร็จสิ้นเรียบร้อยแล้ว'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถทำเครื่องหมายรายการเสร็จสิ้นได้'
            }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/orders/<int:order_id>/complete', methods=['POST'])
def complete_order_endpoint(order_id):
    """ทำเครื่องหมายออเดอร์เสร็จสิ้น"""
    try:
        success = db.complete_order(order_id)
        if success:
            return jsonify({
                'success': True,
                'message': 'ออเดอร์เสร็จสิ้นแล้ว'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่สามารถทำเครื่องหมายออเดอร์เสร็จสิ้นได้'
            }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/staff-request', methods=['POST'])
def staff_request():
    """รับคำขอเรียกพนักงาน"""
    try:
        data = request.get_json()
        table_id = data.get('table_id')
        requests = data.get('requests', [])
        
        if not table_id:
            return jsonify({
                'success': False,
                'error': 'ไม่ได้ระบุหมายเลขโต๊ะ'
            }), 400
            
        if not requests:
            return jsonify({
                'success': False,
                'error': 'ไม่ได้ระบุรายการที่ต้องการ'
            }), 400
        
        # สร้างข้อความแจ้งเตือน
        request_details = []
        for req in requests:
            item_name = req.get('item')
            quantity = req.get('quantity', 0)
            if quantity > 0:
                request_details.append(f"{item_name} {quantity} ชิ้น")
        
        if not request_details:
            return jsonify({
                'success': False,
                'error': 'ไม่มีรายการที่ต้องการ'
            }), 400
        
        message = f"โต๊ะ {table_id} ขอ: {', '.join(request_details)}"
        
        # บันทึกคำขอลงฐานข้อมูล
        notification_data = {
            'table_id': table_id,
            'message': message,
            'timestamp': get_thai_datetime_iso(),
            'type': 'staff_request',
            'is_read': False
        }
        
        # บันทึกการแจ้งเตือนลงฐานข้อมูล
        try:
            db.save_notification(notification_data)
        except Exception as e:
            print(f"Warning: Could not save notification to database: {e}")
        
        return jsonify({
            'success': True,
            'message': 'ส่งคำขอเรียกพนักงานเรียบร้อยแล้ว',
            'notification': notification_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/item-request', methods=['POST'])
def item_request():
    """รับคำขอรายการเพิ่มเติม"""
    try:
        data = request.get_json()
        table_id = data.get('table_id')
        requests = data.get('requests', [])
        
        if not table_id:
            return jsonify({
                'success': False,
                'error': 'ไม่ได้ระบุหมายเลขโต๊ะ'
            }), 400
            
        if not requests:
            return jsonify({
                'success': False,
                'error': 'ไม่ได้ระบุรายการที่ต้องการ'
            }), 400
        
        # สร้างข้อความแจ้งเตือน
        request_details = []
        for req in requests:
            item_name = req.get('item')
            quantity = req.get('quantity', 0)
            if quantity > 0:
                request_details.append(f"{item_name} {quantity} ชิ้น")
        
        if not request_details:
            return jsonify({
                'success': False,
                'error': 'ไม่มีรายการที่ต้องการ'
            }), 400
        
        # สร้างข้อความในรูปแบบที่ admin.js คาดหวัง: "ลูกค้าต้องการภาชนะเพิ่ม|โต๊ะ X|รายการสิ่งของ"
        items_text = ', '.join(request_details)
        message = f"ลูกค้าต้องการภาชนะเพิ่ม|โต๊ะ {table_id}|{items_text}"
        
        # บันทึกคำขอลงฐานข้อมูล
        notification_data = {
            'table_id': table_id,
            'message': message,
            'timestamp': get_thai_datetime_iso(),
            'type': 'item_request',
            'is_read': False
        }
        
        # บันทึกการแจ้งเตือนลงฐานข้อมูล
        try:
            db.save_notification(notification_data)
        except Exception as e:
            print(f"Warning: Could not save notification to database: {e}")
        
        return jsonify({
            'success': True,
            'message': 'ส่งคำขอรายการเพิ่มเติมเรียบร้อยแล้ว',
            'notification': notification_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/order-request', methods=['POST'])
def order_request():
    """รับคำขอการแจ้งเตือนออเดอร์ใหม่"""
    try:
        data = request.get_json()
        table_id = data.get('table_id')
        
        if not table_id:
            return jsonify({
                'success': False,
                'error': 'ไม่ได้ระบุหมายเลขโต๊ะ'
            }), 400
        
        # สร้างข้อความแจ้งเตือนในรูปแบบที่ admin.js คาดหวัง: "ได้รับออเดอร์ใหม่!|โต๊ะ X"
        message = f"ได้รับออเดอร์ใหม่!|โต๊ะ {table_id}"
        
        # บันทึกคำขอลงฐานข้อมูล
        notification_data = {
            'table_id': table_id,
            'message': message,
            'timestamp': get_thai_datetime_iso(),
            'type': 'order_request',
            'is_read': False
        }
        
        # บันทึกการแจ้งเตือนลงฐานข้อมูล
        try:
            db.save_notification(notification_data)
        except Exception as e:
            print(f"Warning: Could not save notification to database: {e}")
        
        return jsonify({
            'success': True,
            'message': 'ส่งการแจ้งเตือนออเดอร์ใหม่เรียบร้อยแล้ว',
            'notification': notification_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/check-bill-request', methods=['POST'])
def check_bill_request():
    """รับคำขอการแจ้งเตือนเช็คบิล"""
    try:
        data = request.get_json()
        table_id = data.get('table_id')
        
        if not table_id:
            return jsonify({
                'success': False,
                'error': 'ไม่ได้ระบุหมายเลขโต๊ะ'
            }), 400
        
        # สร้างข้อความแจ้งเตือนในรูปแบบที่ admin.js คาดหวัง: "ขอเช็คบิล!|โต๊ะ X"
        message = f"ขอเช็คบิล!|โต๊ะ {table_id}"
        
        # บันทึกคำขอลงฐานข้อมูล
        notification_data = {
            'table_id': table_id,
            'message': message,
            'timestamp': get_thai_datetime_iso(),
            'type': 'check_bill_request',
            'is_read': False
        }
        
        # บันทึกการแจ้งเตือนลงฐานข้อมูล
        try:
            db.save_notification(notification_data)
        except Exception as e:
            print(f"Warning: Could not save notification to database: {e}")
        
        return jsonify({
            'success': True,
            'message': 'ส่งการแจ้งเตือนเช็คบิลเรียบร้อยแล้ว',
            'notification': notification_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/notifications', methods=['GET'])
def get_notifications():
    """ดึงการแจ้งเตือนที่ยังไม่ได้อ่าน"""
    try:
        notifications = db.get_unread_notifications()
        return jsonify({
            'success': True,
            'data': notifications
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/notifications/<int:notification_id>/read', methods=['POST'])
def mark_notification_read(notification_id):
    """ทำเครื่องหมายการแจ้งเตือนว่าอ่านแล้ว"""
    try:
        success = db.mark_notification_read(notification_id)
        if success:
            return jsonify({
                'success': True,
                'message': 'ทำเครื่องหมายการแจ้งเตือนว่าอ่านแล้ว'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ไม่พบการแจ้งเตือน'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/orders/<int:order_id>', methods=['GET'])
def get_order_details(order_id):
    """ดึงรายละเอียดออเดอร์ (ทั้งจาก orders และ order_history)"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # ตรวจสอบใน orders table ก่อน
        cursor.execute("""
            SELECT o.order_id, o.table_id, o.session_id, o.status, 
                   o.created_at, o.completed_at, o.total_amount,
                   t.table_name
            FROM orders o
            LEFT JOIN tables t ON o.table_id = t.table_id
            WHERE o.order_id = ?
        """, (order_id,))
        
        order_row = cursor.fetchone()
        
        if order_row:
            # ออเดอร์อยู่ใน orders table (ออเดอร์ปัจจุบัน)
            order_data = {
                'order_id': order_row[0],
                'table_id': order_row[1],
                'session_id': order_row[2],
                'status': order_row[3],
                'created_at': order_row[4],
                'completed_at': order_row[5],
                'total_amount': order_row[6],
                'table_name': order_row[7]
            }
            
            # ดึงรายการอาหารจาก order_items
            cursor.execute("""
                SELECT oi.order_item_id, oi.item_id, oi.quantity, 
                       oi.unit_price, oi.customer_request, oi.status,
                       mi.name, mi.description
                FROM order_items oi
                JOIN menu_items mi ON oi.item_id = mi.item_id
                WHERE oi.order_id = ?
                ORDER BY oi.order_item_id
            """, (order_id,))
            
            items = []
            for item_row in cursor.fetchall():
                items.append({
                    'order_item_id': item_row[0],
                    'item_id': item_row[1],
                    'name': item_row[6],
                    'description': item_row[7],
                    'quantity': item_row[2],
                    'unit_price': item_row[3],
                    'customer_request': item_row[4],
                    'status': item_row[5] or 'pending',
                    'total_price': item_row[2] * item_row[3]
                })
                
        else:
            # ตรวจสอบใน order_history table
            cursor.execute("""
                SELECT oh.order_id, oh.table_id, oh.session_id, oh.status,
                       oh.created_at, oh.completed_at, oh.total_amount,
                       t.table_name
                FROM order_history oh
                LEFT JOIN tables t ON oh.table_id = t.table_id
                WHERE oh.order_id = ?
            """, (order_id,))
            
            history_row = cursor.fetchone()
            
            if not history_row:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': 'ไม่พบออเดอร์'
                }), 404
                
            order_data = {
                'order_id': history_row[0],
                'table_id': history_row[1],
                'session_id': history_row[2],
                'status': history_row[3],
                'created_at': history_row[4],
                'completed_at': history_row[5],
                'total_amount': history_row[6],
                'table_name': history_row[7]
            }
            
            # ดึงรายการอาหารจาก order_history_items
            cursor.execute("""
                SELECT ohi.menu_item_id, ohi.quantity, ohi.price, 
                       ohi.customer_request, ohi.status, mi.name, mi.description
                FROM order_history_items ohi
                LEFT JOIN menu_items mi ON ohi.menu_item_id = mi.item_id
                WHERE ohi.order_id = ?
            """, (order_id,))
            
            items = []
            for item_row in cursor.fetchall():
                items.append({
                    'item_id': item_row[0],
                    'name': item_row[5] or 'รายการที่ถูกลบ',
                    'description': item_row[6] or '',
                    'quantity': item_row[1],
                    'unit_price': item_row[2],
                    'customer_request': item_row[3],
                    'status': item_row[4] or 'completed',
                    'total_price': item_row[1] * item_row[2]
                })
        
        order_data['items'] = items
        conn.close()
        
        return jsonify({
            'success': True,
            'data': order_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/current-orders', methods=['GET'])
def get_current_orders():
    """ดึงสถานะปัจจุบันของคำสั่งซื้อทั้งหมด"""
    try:
        table_id = request.args.get('table_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # สร้าง query สำหรับดึงคำสั่งซื้อปัจจุบัน
        query = """
            SELECT DISTINCT
                o.order_id,
                o.table_id,
                t.table_name as table_name,
                o.session_id,
                o.status,
                o.created_at,
                o.completed_at,
                o.total_amount
            FROM orders o
            LEFT JOIN tables t ON o.table_id = t.table_id
            WHERE 1=1
        """
        
        params = []
        
        # เพิ่มเงื่อนไขการกรอง
        if table_id:
            query += " AND o.table_id = ?"
            params.append(table_id)
            
        if start_date:
            query += " AND DATE(o.created_at) >= ?"
            params.append(start_date)
            
        if end_date:
            query += " AND DATE(o.created_at) <= ?"
            params.append(end_date)
            
        query += " ORDER BY o.created_at DESC"
        
        cursor.execute(query, params)
        orders = []
        
        for row in cursor.fetchall():
            order_data = {
                'order_id': row[0],
                'table_id': row[1],
                'table_name': row[2],
                'session_id': row[3],
                'status': row[4],
                'created_at': row[5],
                'completed_at': row[6],
                'total_amount': row[7]
            }
            
            # ดึงรายการอาหารของแต่ละออเดอร์
            cursor.execute("""
                SELECT 
                    oi.item_id,
                    mi.name as menu_name,
                    oi.quantity,
                    oi.unit_price,
                    oi.customer_request,
                    oi.status
                FROM order_items oi
                LEFT JOIN menu_items mi ON oi.item_id = mi.item_id
                WHERE oi.order_id = ?
            """, (row[0],))
            
            items = []
            for item_row in cursor.fetchall():
                items.append({
                    'menu_item_id': item_row[0],
                    'name': item_row[1],
                    'quantity': item_row[2],
                    'price': item_row[3],
                    'customer_request': item_row[4],
                    'status': item_row[5] if item_row[5] else 'pending',
                    'total_price': item_row[2] * item_row[3]
                })
            
            order_data['items'] = items
            orders.append(order_data)
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': orders
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/order-history', methods=['GET'])
def get_order_history():
    """ดึงประวัติคำสั่งซื้อ"""
    try:
        table_id = request.args.get('table_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # สร้าง query สำหรับดึงประวัติคำสั่งซื้อ
        query = """
            SELECT DISTINCT
                oh.order_id,
                oh.table_id,
                t.table_name as table_name,
                oh.session_id,
                oh.status,
                oh.created_at,
                oh.completed_at,
                oh.total_amount
            FROM order_history oh
            LEFT JOIN tables t ON oh.table_id = t.table_id
            WHERE 1=1
        """
        
        params = []
        
        # เพิ่มเงื่อนไขการกรอง
        if table_id:
            query += " AND oh.table_id = ?"
            params.append(table_id)
            
        if start_date:
            query += " AND DATE(oh.created_at) >= ?"
            params.append(start_date)
            
        if end_date:
            query += " AND DATE(oh.created_at) <= ?"
            params.append(end_date)
            
        query += " ORDER BY oh.created_at DESC"
        
        cursor.execute(query, params)
        orders = []
        
        for row in cursor.fetchall():
            order_data = {
                'order_id': row[0],
                'table_id': row[1],
                'table_name': row[2],
                'session_id': row[3],
                'status': row[4],
                'created_at': row[5],
                'completed_at': row[6],
                'total_amount': row[7]
            }
            
            # ดึงรายการอาหารของแต่ละออเดอร์
            cursor.execute("""
                SELECT 
                    ohi.menu_item_id,
                    mi.name as menu_name,
                    ohi.quantity,
                    ohi.price,
                    ohi.customer_request,
                    ohi.status
                FROM order_history_items ohi
                LEFT JOIN menu_items mi ON ohi.menu_item_id = mi.item_id
                WHERE ohi.order_id = ?
            """, (row[0],))
            
            items = []
            for item_row in cursor.fetchall():
                items.append({
                    'menu_item_id': item_row[0],
                    'name': item_row[1],
                    'quantity': item_row[2],
                    'price': item_row[3],
                    'customer_request': item_row[4],
                    'status': item_row[5] if item_row[5] else 'completed',
                    'total_price': item_row[2] * item_row[3]
                })
            
            order_data['items'] = items
            orders.append(order_data)
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': orders
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/order-history/export', methods=['GET'])
def export_order_history():
    """ส่งออกประวัติคำสั่งซื้อเป็นไฟล์ Excel"""
    try:
        import io
        import pandas as pd
        from datetime import datetime
        
        table_id = request.args.get('table_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # สร้าง query สำหรับดึงข้อมูลส่งออก
        query = """
            SELECT 
                o.order_id as 'หมายเลขออเดอร์',
                t.table_name as 'โต๊ะ',
                o.session_id as 'Session ID',
                o.status as 'สถานะ',
                datetime(o.created_at, 'localtime') as 'วันที่สั่ง',
                datetime(o.completed_at, 'localtime') as 'วันที่เสร็จสิ้น',
                o.total_amount as 'ยอดรวม (บาท)'
            FROM orders o
            LEFT JOIN tables t ON o.table_id = t.table_id
            WHERE 1=1
        """
        
        params = []
        
        if table_id:
            query += " AND o.table_id = ?"
            params.append(table_id)
            
        if start_date:
            query += " AND date(o.created_at) >= ?"
            params.append(start_date)
            
        if end_date:
            query += " AND date(o.created_at) <= ?"
            params.append(end_date)
            
        query += " ORDER BY o.created_at DESC"
        
        cursor.execute(query, params)
        orders_data = cursor.fetchall()
        
        # แปลงข้อมูลเป็น DataFrame
        columns = [description[0] for description in cursor.description]
        df_orders = pd.DataFrame(orders_data, columns=columns)
        
        # ดึงข้อมูลรายการอาหารแยกต่างหาก
        items_query = """
            SELECT 
                oi.order_id as 'หมายเลขออเดอร์',
                mi.name as 'ชื่อเมนู',
                oi.quantity as 'จำนวน',
                oi.unit_price as 'ราคาต่อหน่วย',
                oi.total_price as 'ราคารวม',
                oi.customer_request as 'คำขอพิเศษ',
                oi.status as 'สถานะรายการ'
            FROM order_items oi
            LEFT JOIN menu_items mi ON oi.item_id = mi.item_id
            WHERE oi.order_id IN (
                SELECT o.order_id FROM orders o
                LEFT JOIN tables t ON o.table_id = t.table_id
                WHERE 1=1
        """
        
        if table_id:
            items_query += " AND o.table_id = ?"
            
        if start_date:
            items_query += " AND date(o.created_at) >= ?"
            
        if end_date:
            items_query += " AND date(o.created_at) <= ?"
            
        items_query += ")"
        
        cursor.execute(items_query, params)
        items_data = cursor.fetchall()
        
        columns_items = [description[0] for description in cursor.description]
        df_items = pd.DataFrame(items_data, columns=columns_items)
        
        conn.close()
        
        # สร้างไฟล์ Excel พร้อมการจัดรูปแบบ
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # เขียนข้อมูลออเดอร์
            df_orders.to_excel(writer, sheet_name='ออเดอร์', index=False, startrow=1)
            
            # เขียนข้อมูลรายการอาหาร
            df_items.to_excel(writer, sheet_name='รายการอาหาร', index=False, startrow=1)
            
            # จัดรูปแบบชีต ออเดอร์
            ws_orders = writer.sheets['ออเดอร์']
            
            # จัดรูปแบบ header ของตาราง
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # เพิ่มหัวข้อหลัก
            ws_orders['A1'] = 'รายงานประวัติการสั่งอาหาร'
            ws_orders['A1'].font = Font(size=16, bold=True)
            ws_orders.merge_cells('A1:G1')
            
            # คำนวณสรุปยอด
            total_orders = len(df_orders)
            total_amount = df_orders['ยอดรวม (บาท)'].sum() if not df_orders.empty else 0
            avg_amount = df_orders['ยอดรวม (บาท)'].mean() if not df_orders.empty else 0
            
            # เพิ่มสรุปยอดในชีตออเดอร์
            summary_row = len(df_orders) + 4
            ws_orders[f'A{summary_row}'] = 'สรุปยอด'
            ws_orders[f'A{summary_row}'].font = Font(bold=True, size=14)
            ws_orders[f'A{summary_row + 1}'] = f'จำนวนออเดอร์ทั้งหมด: {total_orders} ออเดอร์'
            ws_orders[f'A{summary_row + 2}'] = f'ยอดขายรวม: {total_amount:,.2f} บาท'
            ws_orders[f'A{summary_row + 3}'] = f'ยอดขายเฉลี่ยต่อออเดอร์: {avg_amount:,.2f} บาท'
            
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            center_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # จัดรูปแบบ header ชีตออเดอร์
            for col in range(1, len(df_orders.columns) + 1):
                cell = ws_orders.cell(row=2, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # จัดรูปแบบข้อมูลในตาราง
            for row in range(3, len(df_orders) + 3):
                for col in range(1, len(df_orders.columns) + 1):
                    cell = ws_orders.cell(row=row, column=col)
                    cell.border = thin_border
                    if col in [7]:  # คอลัมน์ยอดรวม
                        cell.alignment = Alignment(horizontal='right')
            
            # ปรับความกว้างคอลัมน์
            column_widths = [15, 12, 20, 12, 20, 20, 15]
            column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
            for i, width in enumerate(column_widths):
                ws_orders.column_dimensions[column_letters[i]].width = width
            
            # จัดรูปแบบชีต รายการอาหาร
            ws_items = writer.sheets['รายการอาหาร']
            
            # เพิ่มหัวข้อหลัก
            ws_items['A1'] = 'รายละเอียดรายการอาหาร'
            ws_items['A1'].font = Font(size=16, bold=True)
            ws_items.merge_cells('A1:G1')
            
            # จัดรูปแบบ header ชีตรายการอาหาร
            for col in range(1, len(df_items.columns) + 1):
                cell = ws_items.cell(row=2, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # จัดรูปแบบข้อมูลในตาราง
            for row in range(3, len(df_items) + 3):
                for col in range(1, len(df_items.columns) + 1):
                    cell = ws_items.cell(row=row, column=col)
                    cell.border = thin_border
                    if col in [4, 5]:  # คอลัมน์ราคา
                        cell.alignment = Alignment(horizontal='right')
                    elif col == 3:  # คอลัมน์จำนวน
                        cell.alignment = Alignment(horizontal='center')
            
            # ปรับความกว้างคอลัมน์สำหรับชีตรายการอาหาร
            item_column_widths = [15, 25, 10, 15, 15, 30, 15]
            for i, width in enumerate(item_column_widths):
                ws_items.column_dimensions[column_letters[i]].width = width
            
            # เพิ่มสรุปยอดในชีตรายการอาหาร
            if not df_items.empty:
                total_items = df_items['จำนวน'].sum()
                total_item_amount = df_items['ราคารวม'].sum()
                
                item_summary_row = len(df_items) + 4
                ws_items[f'A{item_summary_row}'] = 'สรุปยอดรายการอาหาร'
                ws_items[f'A{item_summary_row}'].font = Font(bold=True, size=14)
                ws_items[f'A{item_summary_row + 1}'] = f'จำนวนรายการทั้งหมด: {total_items} รายการ'
                ws_items[f'A{item_summary_row + 2}'] = f'ยอดขายรวม: {total_item_amount:,.2f} บาท'
        
        output.seek(0)
        
        # สร้างชื่อไฟล์
        filename = f"order_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/css/<path:filename>')
def css_files(filename):
    return send_from_directory('../frontend/css', filename)

@app.route('/js/<path:filename>')
def js_files(filename):
    return send_from_directory('../frontend/js', filename)

@app.route('/assets/<path:filename>')
def asset_files(filename):
    return send_from_directory('../frontend/assets', filename)

@app.route('/images/<path:filename>')
def image_files(filename):
    return send_from_directory('../frontend/images', filename)

def allowed_file(filename):
    """ตรวจสอบว่าไฟล์มีนามสกุลที่อนุญาตหรือไม่"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/api/upload/menu-image', methods=['POST'])
def upload_menu_image():
    """อัปโหลดรูปภาพเมนูอาหารและปรับให้เป็นอัตราส่วน 1:1 โดยอัตโนมัติ
    รับพารามิเตอร์ menu_id และ menu_name เพื่อตั้งชื่อไฟล์ตามชื่อเมนู
    และทับไฟล์เดิมถ้ามีการอัปโหลดใหม่"""
    try:
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'ไม่พบไฟล์ในคำขอ'
            }), 400
            
        file = request.files['file']
        menu_id = request.form.get('menu_id')
        menu_name = request.form.get('menu_name')
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'ไม่ได้เลือกไฟล์'
            }), 400
        
        if not menu_name:
            return jsonify({
                'success': False,
                'error': 'ไม่ได้ระบุชื่อเมนู'
            }), 400
            
        if file and allowed_file(file.filename):
            # สร้างชื่อไฟล์จากชื่อเมนู
            extension = file.filename.rsplit('.', 1)[1].lower()
            safe_menu_name = secure_filename(menu_name)
            
            # ถ้ามี menu_id ให้ใช้ในชื่อไฟล์ด้วย
            if menu_id:
                unique_filename = f"menu_{menu_id}_{safe_menu_name}.{extension}"
            else:
                # ถ้าไม่มี menu_id ให้ใช้เวลาปัจจุบันแทน
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                unique_filename = f"{safe_menu_name}_{timestamp}.{extension}"
                
            # ตรวจสอบว่ามีไฟล์เดิมหรือไม่ (กรณีมี menu_id)
            if menu_id:
                # ค้นหาไฟล์เดิมที่มีรูปแบบ menu_{menu_id}_*.* เพื่อลบทิ้ง
                menu_image_pattern = os.path.join(UPLOAD_FOLDER, f"menu_{menu_id}_*.*")
                old_files = glob.glob(menu_image_pattern)
                for old_file in old_files:
                    try:
                        os.remove(old_file)
                    except Exception as e:
                        print(f"ไม่สามารถลบไฟล์เดิม {old_file}: {e}")
            
            # บันทึกไฟล์ชั่วคราว
            temp_file_path = os.path.join(UPLOAD_FOLDER, f"temp_{unique_filename}")
            file.save(temp_file_path)
            
            # ปรับภาพให้เป็นอัตราส่วน 1:1
            with Image.open(temp_file_path) as img:
                # หาขนาดที่ใหญ่ที่สุดระหว่างความกว้างและความสูง
                width, height = img.size
                size = max(width, height)
                
                # สร้างภาพใหม่ที่เป็นสี่เหลี่ยมจัตุรัส (1:1) โดยใช้พื้นหลังสีขาว
                square_img = Image.new('RGBA', (size, size), (255, 255, 255, 0))
                
                # วางภาพต้นฉบับตรงกลางของภาพใหม่
                paste_x = (size - width) // 2
                paste_y = (size - height) // 2
                square_img.paste(img, (paste_x, paste_y))
                
                # บันทึกภาพที่ปรับแล้ว
                file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
                square_img.save(file_path)
                
                # ลบไฟล์ชั่วคราว
                os.remove(temp_file_path)
            
            # สร้าง URL สำหรับเข้าถึงไฟล์
            image_url = f"/images/menu/{unique_filename}"
            
            # อัปเดตฐานข้อมูลถ้ามี menu_id
            if menu_id:
                try:
                    # ดึงข้อมูลเมนูปัจจุบัน
                    menu_item = db.get_menu_item(int(menu_id))
                    if menu_item:
                        # อัปเดตเฉพาะ image_url
                        success = db.update_menu_item(
                            int(menu_id),
                            menu_item['name'],
                            menu_item['price'],
                            menu_item['category_id'],
                            menu_item.get('description', ''),
                            image_url,
                            menu_item.get('is_available', True),
                            menu_item.get('preparation_time', 15)
                        )
                        if not success:
                            print(f"ไม่สามารถอัปเดต image_url ในฐานข้อมูลสำหรับเมนู ID: {menu_id}")
                except Exception as e:
                    print(f"เกิดข้อผิดพลาดในการอัปเดตฐานข้อมูล: {e}")
            
            return jsonify({
                'success': True,
                'image_url': image_url,
                'menu_id': menu_id,
                'message': 'อัปโหลดรูปภาพสำเร็จและบันทึกลงฐานข้อมูลแล้ว'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'ประเภทไฟล์ไม่ได้รับอนุญาต'
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/total-price-preview.html')
def total_price_preview():
    """หน้าตัวอย่างการแสดงยอดรวม"""
    return render_template('total-price-preview.html')

@app.route('/api/dashboard-data', methods=['GET'])
def get_dashboard_data():
    """API สำหรับข้อมูล Dashboard"""
    try:
        start_date = request.args.get('start')
        end_date = request.args.get('end')
        range_param = request.args.get('range')
        
        # Import datetime modules
        from datetime import date, timedelta
        
        # จัดการ range parameter
        if range_param:
            today_date = date.today()
            if range_param == 'today':
                start_date = today_date.strftime('%Y-%m-%d')
                end_date = today_date.strftime('%Y-%m-%d')
            elif range_param == 'week':
                start_date_obj = today_date - timedelta(days=6)  # 7 วันย้อนหลัง
                start_date = start_date_obj.strftime('%Y-%m-%d')
                end_date = today_date.strftime('%Y-%m-%d')
            elif range_param == 'month':
                start_date_obj = today_date - timedelta(days=29)  # 30 วันย้อนหลัง
                start_date = start_date_obj.strftime('%Y-%m-%d')
                end_date = today_date.strftime('%Y-%m-%d')
        
        # ถ้าไม่มีการระบุวันที่ ให้ใช้ช่วง 7 วันย้อนหลัง (รวมวันนี้)
        if not start_date or not end_date:
            today_date = date.today()
            start_date_obj = today_date - timedelta(days=6)  # 7 วันย้อนหลัง
            start_date = start_date_obj.strftime('%Y-%m-%d')
            end_date = today_date.strftime('%Y-%m-%d')
        
        print(f"[DEBUG] Dashboard API called with range={range_param}, start={start_date}, end={end_date}")
        
        # ดึงข้อมูลจากฐานข้อมูล
        orders = db.get_orders_by_date_range(start_date, end_date)
        print(f"[DEBUG] Orders from get_orders_by_date_range({start_date}, {end_date}): {len(orders)}")
        for i, order in enumerate(orders[:3]):  # แสดงแค่ 3 รายการแรก
            print(f"[DEBUG] Order {i+1}: ID={order.get('order_id')}, status={order.get('status')}, amount={order.get('total_amount')}")
        
        # คำนวณข้อมูลสำหรับ dashboard
        period_sales = 0
        today_sales = 0
        week_sales = 0
        month_sales = 0
        total_customers = 0
        daily_sales = {}
        category_sales = {}
        top_items = []
        
        # วันนี้
        today_date = date.today()
        today_str = today_date.strftime('%Y-%m-%d')
        
        # สัปดาห์นี้ (จันทร์ถึงอาทิตย์)
        days_since_monday = today_date.weekday()  # 0=จันทร์, 6=อาทิตย์
        week_start = (today_date - timedelta(days=days_since_monday)).strftime('%Y-%m-%d')
        week_end = (today_date + timedelta(days=6-days_since_monday)).strftime('%Y-%m-%d')
        
        # เดือนนี้
        month_start = today_date.replace(day=1).strftime('%Y-%m-%d')
        
        # ดึงข้อมูลสำหรับแต่ละช่วงเวลา
        today_orders = db.get_orders_by_date_range(today_str, today_str)
        week_orders = db.get_orders_by_date_range(week_start, week_end)
        month_orders = db.get_orders_by_date_range(month_start, today_str)
        
        # คำนวณยอดขาย (รวมทุกสถานะยกเว้น rejected)
        for order in today_orders:
            if order.get('status') != 'rejected':
                today_sales += order.get('total_amount', 0)
                
        for order in week_orders:
            if order.get('status') != 'rejected':
                week_sales += order.get('total_amount', 0)
                
        for order in month_orders:
            if order.get('status') != 'rejected':
                month_sales += order.get('total_amount', 0)
        

        
        # ยอดขายในช่วงที่เลือก
        for order in orders:
            if order.get('status') != 'rejected':
                period_sales += order.get('total_amount', 0)
                total_customers += 1
                
                # จัดกลุ่มตามวันที่สำหรับ chart
                order_date = order.get('created_at', '')[:10]  # YYYY-MM-DD
                if order_date not in daily_sales:
                    daily_sales[order_date] = {'sales': 0, 'orders': 0}
                daily_sales[order_date]['sales'] += order.get('total_amount', 0)
                daily_sales[order_date]['orders'] += 1
        
        # สร้างข้อมูล mock สำหรับ category sales
        categories = db.get_menu_categories()
        for cat in categories:
            category_sales[cat['name']] = 0
        
        # ข้อมูล top items จากข้อมูลจริงในช่วงเวลาที่เลือก
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT mi.name, SUM(oi.quantity) as total_quantity, SUM(oi.total_price) as total_sales
                FROM order_items oi
                JOIN menu_items mi ON oi.item_id = mi.item_id
                JOIN orders o ON oi.order_id = o.order_id
                WHERE o.status != 'rejected'
                AND DATE(o.created_at) BETWEEN ? AND ?
                GROUP BY mi.item_id, mi.name
                ORDER BY total_quantity DESC
                LIMIT 5
            """, (start_date, end_date))
            
            for row in cursor.fetchall():
                top_items.append({
                    'name': row['name'],
                    'quantity': int(row['total_quantity']),
                    'sales': float(row['total_sales']) if row['total_sales'] else 0
                })
            
            conn.close()
        except Exception as e:
            print(f"[ERROR] Failed to get top items: {e}")
            # Fallback to empty list if query fails
            top_items = []
        
        response_data = {
            'periodSales': period_sales,
            'todaySales': today_sales,
            'weekSales': week_sales,
            'currentWeekSales': week_sales,
            'monthSales': month_sales,
            'totalCustomers': total_customers,
            'dailySales': daily_sales,
            'categorySales': category_sales,
            'topItems': top_items,
            'monthlyTrend': []
        }
        
        print(f"[DEBUG] ===== FINAL RESPONSE ===== {response_data}")
        print(f"[DEBUG] ===== DASHBOARD API COMPLETED =====")
        return jsonify({
            'success': True,
            'data': response_data
        })
        
    except Exception as e:
        print(f"[ERROR] Dashboard API error: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'data': {
                'periodSales': 0,
                'todaySales': 0,
                'weekSales': 0,
                'currentWeekSales': 0,
                'monthSales': 0,
                'totalCustomers': 0,
                'dailySales': {},
                'categorySales': {},
                'topItems': [],
                'monthlyTrend': []
            }
        }), 500

@app.route('/api/sales-summary', methods=['GET'])
def get_sales_summary():
    """API สำหรับข้อมูลสรุปยอดขาย"""
    try:
        print("[DEBUG SALES] Sales summary API called")
        
        from datetime import date, timedelta
        today = date.today()
        today_str = today.strftime('%Y-%m-%d')
        print(f"[DEBUG SALES] Today date: {today_str}")
        
        # สัปดาห์นี้ (จันทร์ถึงอาทิตย์)
        days_since_monday = today.weekday()  # 0=จันทร์, 6=อาทิตย์
        week_start = (today - timedelta(days=days_since_monday)).strftime('%Y-%m-%d')
        week_end = (today + timedelta(days=6-days_since_monday)).strftime('%Y-%m-%d')
        print(f"[DEBUG SALES] Week range: {week_start} to {week_end}")
        
        # เดือนนี้
        month_start = today.replace(day=1).strftime('%Y-%m-%d')
        print(f"[DEBUG SALES] Month range: {month_start} to {today_str}")
        
        # ดึงข้อมูลจากฐานข้อมูล
        today_orders = db.get_orders_by_date_range(today_str, today_str)
        print(f"[DEBUG SALES] Today orders count: {len(today_orders)}")
        
        week_orders = db.get_orders_by_date_range(week_start, week_end)
        print(f"[DEBUG SALES] Week orders count: {len(week_orders)}")
        
        month_orders = db.get_orders_by_date_range(month_start, today_str)
        print(f"[DEBUG SALES] Month orders count: {len(month_orders)}")
        
        # ใช้วันที่เริ่มต้นที่เก่ามากเพื่อดึงออเดอร์ทั้งหมด
        all_orders = db.get_orders_by_date_range('2020-01-01', today_str)
        print(f"[DEBUG SALES] All orders count: {len(all_orders)}")
        
        # คำนวณยอดขาย (รวมทุกสถานะยกเว้น rejected)
        today_completed = [o for o in today_orders if o.get('status') != 'rejected']
        today_total = sum(order.get('total_amount', 0) for order in today_completed)
        print(f"[DEBUG SALES] Today non-rejected orders: {len(today_completed)}, total: {today_total}")
        
        week_completed = [o for o in week_orders if o.get('status') != 'rejected']
        week_total = sum(order.get('total_amount', 0) for order in week_completed)
        print(f"[DEBUG SALES] Week non-rejected orders: {len(week_completed)}, total: {week_total}")
        
        month_completed = [o for o in month_orders if o.get('status') != 'rejected']
        month_total = sum(order.get('total_amount', 0) for order in month_completed)
        print(f"[DEBUG SALES] Month non-rejected orders: {len(month_completed)}, total: {month_total}")
        
        all_completed = [o for o in all_orders if o.get('status') != 'rejected']
        total_total = sum(order.get('total_amount', 0) for order in all_completed)
        print(f"[DEBUG SALES] All non-rejected orders: {len(all_completed)}, total: {total_total}")
        
        # นับจำนวน orders
        today_orders_count = len([o for o in today_orders if o.get('status') != 'rejected'])
        week_orders_count = len([o for o in week_orders if o.get('status') != 'rejected'])
        month_orders_count = len([o for o in month_orders if o.get('status') != 'rejected'])
        
        # นับจำนวน sessions (ใช้ table_id ที่ unique)
        today_sessions = len(set(order.get('table_id') for order in today_orders if order.get('status') != 'rejected'))
        
        response_data = {
            'success': True,
            'data': {
                'today': {
                    'total': today_total,
                    'orders': today_orders_count,
                    'sessions': today_sessions
                },
                'week': {
                    'total': week_total,
                    'orders': week_orders_count
                },
                'month': {
                    'total': month_total,
                    'orders': month_orders_count
                },
                'total': {
                    'total': total_total
                }
            }
        }
        
        print(f"[DEBUG] Sales summary response: {response_data}")
        return jsonify(response_data)
        
    except Exception as e:
        print(f"[ERROR] Sales summary API error: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'data': {
                'today': {'total': 0, 'orders': 0, 'sessions': 0},
                'week': {'total': 0, 'orders': 0},
                'month': {'total': 0, 'orders': 0},
                'total': {'total': 0}
            }
        }), 500

@app.route('/api/sales-summary/monthly', methods=['GET'])
def get_monthly_sales_summary():
    """API สำหรับข้อมูลสรุปยอดขายรายเดือน"""
    try:
        year = request.args.get('year')
        month = request.args.get('month')
        
        print(f"[DEBUG] Monthly sales summary API called with year={year}, month={month}")
        
        if not year or not month:
            return jsonify({'success': False, 'error': 'Missing year or month parameter'}), 400
        
        # สร้างช่วงวันที่สำหรับเดือนที่ระบุ
        from datetime import date
        import calendar
        
        start_date = f"{year}-{month.zfill(2)}-01"
        last_day = calendar.monthrange(int(year), int(month))[1]
        end_date = f"{year}-{month.zfill(2)}-{last_day:02d}"
        
        # ดึงข้อมูลจากฐานข้อมูล
        orders = db.get_orders_by_date_range(start_date, end_date)
        
        # คำนวณยอดขาย (รวมทุกสถานะยกเว้น rejected)
        total = sum(order.get('total_amount', 0) for order in orders if order.get('status') != 'rejected')
        sessions = len(set(order.get('table_id') for order in orders if order.get('status') != 'rejected'))
        
        response_data = {
            'success': True,
            'total': total,
            'sessions': sessions
        }
        
        print(f"[DEBUG] Monthly sales summary response: {response_data}")
        return jsonify(response_data)
        
    except Exception as e:
        print(f"[ERROR] Monthly sales summary API error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sales-summary/custom', methods=['GET'])
def get_custom_sales_summary():
    """API สำหรับข้อมูลสรุปยอดขายตามช่วงวันที่กำหนดเอง"""
    try:
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')
        
        print(f"[DEBUG] Custom sales summary API called with startDate={start_date}, endDate={end_date}")
        
        if not start_date or not end_date:
            return jsonify({'success': False, 'error': 'Missing startDate or endDate parameter'}), 400
        
        # ดึงข้อมูลจากฐานข้อมูล
        orders = db.get_orders_by_date_range(start_date, end_date)
        
        # คำนวณยอดขาย (รวมทุกสถานะยกเว้น rejected)
        total = sum(order.get('total_amount', 0) for order in orders if order.get('status') != 'rejected')
        sessions = len(set(order.get('table_id') for order in orders if order.get('status') != 'rejected'))
        
        response_data = {
            'success': True,
            'total': total,
            'sessions': sessions
        }
        
        print(f"[DEBUG] Custom sales summary response: {response_data}")
        return jsonify(response_data)
        
    except Exception as e:
        print(f"[ERROR] Custom sales summary API error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/sales-chart', methods=['GET'])
def get_sales_chart():
    """ดึงข้อมูลกราฟยอดขาย"""
    try:
        days = request.args.get('days', 7, type=int)
        
        # ดึงข้อมูลยอดขายตามวัน
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT DATE(created_at) as date, SUM(total_amount) as total
            FROM orders 
            WHERE status != 'rejected' 
            AND DATE(created_at) >= DATE('now', '-{} days')
            GROUP BY DATE(created_at)
            ORDER BY date
        """.format(days))
        
        sales_data = []
        for row in cursor.fetchall():
            sales_data.append({
                'date': row['date'],
                'total': float(row['total']) if row['total'] else 0
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': sales_data
        })
        
    except Exception as e:
        print(f"[ERROR] Sales chart API error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/top-items', methods=['GET'])
def get_top_items():
    """ดึงข้อมูลเมนูขายดี"""
    try:
        limit = request.args.get('limit', 5, type=int)
        
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT mi.name, SUM(oi.quantity) as total_quantity, SUM(oi.total_price) as total_sales
            FROM order_items oi
            JOIN menu_items mi ON oi.item_id = mi.item_id
            JOIN orders o ON oi.order_id = o.order_id
            WHERE o.status != 'rejected'
            GROUP BY mi.item_id, mi.name
            ORDER BY total_quantity DESC
            LIMIT ?
        """, (limit,))
        
        top_items = []
        for row in cursor.fetchall():
            top_items.append({
                'name': row['name'],
                'quantity': int(row['total_quantity']),
                'sales': float(row['total_sales']) if row['total_sales'] else 0
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': top_items
        })
        
    except Exception as e:
        print(f"[ERROR] Top items API error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/category-chart', methods=['GET'])
def get_category_chart():
    """ดึงข้อมูลกราฟยอดขายตามหมวดหมู่"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT mc.name as category_name, SUM(oi.total_price) as total_sales
            FROM order_items oi
            JOIN menu_items mi ON oi.item_id = mi.item_id
            JOIN menu_categories mc ON mi.category_id = mc.category_id
            JOIN orders o ON oi.order_id = o.order_id
            WHERE o.status != 'rejected'
            GROUP BY mc.category_id, mc.name
            ORDER BY total_sales DESC
        """)
        
        category_data = []
        for row in cursor.fetchall():
            category_data.append({
                'category': row['category_name'],
                'sales': float(row['total_sales']) if row['total_sales'] else 0
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': category_data
        })
        
    except Exception as e:
        print(f"[ERROR] Category chart API error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/restaurant-info', methods=['GET'])
def get_restaurant_info():
    """ดึงข้อมูลร้านอาหาร"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # ดึงข้อมูลร้านอาหารจากฐานข้อมูล
        cursor.execute("SELECT config_value FROM system_config WHERE config_key = 'restaurant_name'")
        name_result = cursor.fetchone()
        
        cursor.execute("SELECT config_value FROM system_config WHERE config_key = 'restaurant_address'")
        address_result = cursor.fetchone()
        
        cursor.execute("SELECT config_value FROM system_config WHERE config_key = 'restaurant_phone'")
        phone_result = cursor.fetchone()
        
        conn.close()
        
        # ใช้ค่าเริ่มต้นหากไม่มีข้อมูลในฐานข้อมูล
        restaurant_info = {
            'name': name_result['config_value'] if name_result else 'ร้านอาหาร A-FOOD',
            'address': address_result['config_value'] if address_result else 'สงขลา หาดใหญ่',
            'phone': phone_result['config_value'] if phone_result else '02-xxx-xxxx'
        }
        
        return jsonify({
            'success': True,
            'data': restaurant_info
        })
        
    except Exception as e:
        print(f"[ERROR] Restaurant info API error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/restaurant-info', methods=['POST'])
def save_restaurant_info():
    """บันทึกข้อมูลร้านอาหาร"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': 'ไม่มีข้อมูลที่ส่งมา'}), 400
        
        # ตรวจสอบข้อมูลที่จำเป็น
        restaurant_name = data.get('restaurant_name', '').strip()
        restaurant_address = data.get('restaurant_address', '').strip()
        restaurant_phone = data.get('restaurant_phone', '').strip()
        
        if not restaurant_name:
            return jsonify({'success': False, 'error': 'กรุณากรอกชื่อร้าน'}), 400
        
        # บันทึกข้อมูลลงฐานข้อมูล
        db.save_setting('restaurant_name', restaurant_name)
        db.save_setting('restaurant_address', restaurant_address)
        db.save_setting('restaurant_phone', restaurant_phone)
        
        return jsonify({
            'success': True,
            'message': 'บันทึกข้อมูลร้านอาหารเรียบร้อยแล้ว',
            'data': {
                'restaurant_name': restaurant_name,
                'restaurant_address': restaurant_address,
                'restaurant_phone': restaurant_phone
            }
        })
        
    except Exception as e:
        print(f"[ERROR] Save restaurant info API error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== LOGIN API ====================

@app.route('/api/login', methods=['POST'])
def login():
    """API สำหรับการ login"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'ไม่มีข้อมูลที่ส่งมา'}), 400
        
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        # ตรวจสอบ username และ password
        if username == 'admin' and password == 'korjud':
            # สร้าง session
            session['logged_in'] = True
            session['username'] = username
            session['login_time'] = datetime.now().isoformat()
            
            print(f"[LOGIN] User '{username}' logged in successfully")
            
            return jsonify({
                'success': True,
                'message': 'เข้าสู่ระบบสำเร็จ',
                'user': username
            })
        else:
            print(f"[LOGIN] Failed login attempt for username: '{username}'")
            return jsonify({
                'success': False,
                'message': 'ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง'
            }), 401
            
    except Exception as e:
        print(f"[ERROR] Login API error: {e}")
        return jsonify({'success': False, 'message': 'เกิดข้อผิดพลาดในระบบ'}), 500

@app.route('/api/logout', methods=['POST'])
def logout():
    """API สำหรับการ logout"""
    try:
        username = session.get('username', 'Unknown')
        session.clear()
        
        print(f"[LOGOUT] User '{username}' logged out")
        
        return jsonify({
            'success': True,
            'message': 'ออกจากระบบสำเร็จ'
        })
        
    except Exception as e:
        print(f"[ERROR] Logout API error: {e}")
        return jsonify({'success': False, 'message': 'เกิดข้อผิดพลาดในระบบ'}), 500

@app.route('/api/check-auth', methods=['GET'])
def check_auth():
    """ตรวจสอบสถานะการ login"""
    try:
        if session.get('logged_in'):
            return jsonify({
                'success': True,
                'logged_in': True,
                'username': session.get('username'),
                'login_time': session.get('login_time')
            })
        else:
            return jsonify({
                'success': True,
                'logged_in': False
            })
            
    except Exception as e:
        print(f"[ERROR] Check auth API error: {e}")
        return jsonify({'success': False, 'message': 'เกิดข้อผิดพลาดในระบบ'}), 500

if __name__ == '__main__':
    print("\n" + "="*50)
    print("🍽️ เริ่มต้นระบบ POS ร้านอาหาร")
    print("📱 ลูกค้าสามารถเข้าถึงได้ที่: http://localhost:5000")
    print("🖥️ ระบบหลังบ้าน: เปิดโปรแกรม Desktop App")
    print("\n🚀 กำลังเริ่มต้น Flask Server...")
    
    app.run(
        host='0.0.0.0',  # เปิดให้เครื่องอื่นเข้าถึงได้
        port=5000,
        debug=True
    )